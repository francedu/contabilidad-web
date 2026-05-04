from __future__ import annotations

import os
import re
import sqlite3
import subprocess
import smtplib
from email.message import EmailMessage
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import csv
import shutil
import secrets
from io import BytesIO, StringIO

from flask import Flask, flash, g, redirect, render_template, request, send_file, url_for
from flask_login import LoginManager, UserMixin, current_user, login_required, login_user, logout_user
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except Exception:
    psycopg2 = None
    RealDictCursor = None

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / 'instance' / 'contabilidad_curso.db'
INSTANCE_DIR = BASE_DIR / 'instance'
BACKUP_DIR = BASE_DIR / 'backups'
INSTANCE_DIR.mkdir(exist_ok=True)
BACKUP_DIR.mkdir(exist_ok=True)

APP_NAME = 'ContaCurso'
APP_TAGLINE = 'Plataforma de administración de cuotas y fondos escolares'
SCHOOL_NAME = 'Escuela Las Mercedes'
SCHOOL_LOCATION = 'María Pinto'
GLOBAL_ROLES = ('admin',)
COURSE_ROLES = ('presidente', 'tesorero', 'secretario', 'apoderado')
WRITE_COURSE_ROLES = ('presidente', 'tesorero', 'secretario')
ALLOWED_ROLES = GLOBAL_ROLES + COURSE_ROLES + ('solo_lectura',)
PLANES_COLEGIO = ('basico', 'pro')
ESTADOS_SUSCRIPCION = ('activo', 'vencido')
PREDEFINED_COURSES = [
    'Prekínder',
    'Kínder',
    '1° Básico',
    '2° Básico',
    '3° Básico',
    '4° Básico',
    '5° Básico',
    '6° Básico',
    '7° Básico',
    '8° Básico',
]

COURSE_NORMALIZATION_MAP = {
    'Prekínder': ['prekinder', 'pre kinder', 'pre-kinder', 'pre kínder', 'prekínder'],
    'Kínder': ['kinder', 'kínder'],
    '1° Básico': ['1', '1 basico', '1 básico', '1ro', '1ro basico', '1ro básico', 'primero', 'primero basico', 'primero básico'],
    '2° Básico': ['2', '2 basico', '2 básico', '2do', '2do basico', '2do básico', 'segundo', 'segundo basico', 'segundo básico'],
    '3° Básico': ['3', '3 basico', '3 básico', '3ro', '3ro basico', '3ro básico', 'tercero', 'tercero basico', 'tercero básico'],
    '4° Básico': ['4', '4 basico', '4 básico', '4to', '4to basico', '4to básico', 'cuarto', 'cuarto basico', 'cuarto básico'],
    '5° Básico': ['5', '5 basico', '5 básico', '5to', '5to basico', '5to básico', 'quinto', 'quinto basico', 'quinto básico'],
    '6° Básico': ['6', '6 basico', '6 básico', '6to', '6to basico', '6to básico', 'sexto', 'sexto basico', 'sexto básico'],
    '7° Básico': ['7', '7 basico', '7 básico', '7mo', '7mo basico', '7mo básico', 'séptimo', 'septimo', 'séptimo basico', 'séptimo básico', 'septimo basico', 'septimo básico'],
    '8° Básico': ['8', '8 basico', '8 básico', '8vo', '8vo basico', '8vo básico', 'octavo', 'octavo basico', 'octavo básico'],
}


def normalize_course_value(curso: str | None) -> str:
    value = (curso or '').strip().lower()
    replacements = str.maketrans({
        'á': 'a',
        'é': 'e',
        'í': 'i',
        'ó': 'o',
        'ú': 'u',
    })
    return value.translate(replacements)


def standardize_course_name(curso: str | None) -> str | None:
    value = normalize_course_value(curso)
    if not value:
        return None
    for canonical, aliases in COURSE_NORMALIZATION_MAP.items():
        normalized_aliases = {normalize_course_value(alias) for alias in aliases}
        normalized_aliases.add(normalize_course_value(canonical))
        if value in normalized_aliases:
            return canonical
    return (curso or '').strip() or None


def normalize_courses_in_db(db: 'DBAdapter') -> None:
    tables = ['usuarios', 'alumnos', 'movimientos', 'actividades']
    for table in tables:
        try:
            rows = db.fetchall(f"SELECT id, curso FROM {table} WHERE curso IS NOT NULL AND trim(curso) <> ''")
        except Exception:
            db.rollback()
            continue
        changed = False
        for row in rows:
            current_value = (row['curso'] or '').strip()
            canonical = standardize_course_name(current_value)
            if canonical and canonical != current_value:
                db.execute(f"UPDATE {table} SET curso = ? WHERE id = ?", (canonical, row['id']))
                changed = True
        if changed:
            try:
                db.commit()
            except Exception:
                db.rollback()


class User(UserMixin):
    def __init__(self, row: Any):
        self.id = str(row['id'])
        self.username = row['username']
        self.password_hash = row['password_hash']
        self.role = row['role']
        self.nombre = row['nombre']
        self.email = row['email'] if 'email' in row.keys() else None
        self.curso = row['curso'] if 'curso' in row.keys() else None
        self.colegio_id = row['colegio_id'] if 'colegio_id' in row.keys() else None
        self.activo = bool(row['activo'])

    def is_admin_global(self) -> bool:
        return self.role in ('admin_global', 'admin')

    def is_admin(self) -> bool:
        return self.is_admin_global()

    def can_edit(self) -> bool:
        return self.is_admin_global() or self.role in WRITE_COURSE_ROLES

    def can_delete(self) -> bool:
        return self.is_admin_global() or self.role in ('presidente', 'tesorero')


class DBAdapter:
    def __init__(self, url: str):
        self.url = url
        self.kind = 'postgres' if url.startswith('postgresql://') or url.startswith('postgres://') else 'sqlite'
        self.conn = self._connect()

    def _connect(self):
        if self.kind == 'sqlite':
            conn = sqlite3.connect(self.url)
            conn.row_factory = sqlite3.Row
            conn.execute('PRAGMA foreign_keys = ON')
            return conn
        if psycopg2 is None:
            raise RuntimeError('psycopg2-binary no está instalado. Ejecuta: python3 -m pip install psycopg2-binary')
        conn = psycopg2.connect(self.url, cursor_factory=RealDictCursor)
        conn.autocommit = False
        return conn

    def close(self):
        self.conn.close()

    def commit(self):
        self.conn.commit()

    def rollback(self):
        self.conn.rollback()

    def cursor(self):
        return self.conn.cursor()

    def _convert_sql(self, sql: str) -> str:
        if self.kind == 'sqlite':
            return sql
        return re.sub(r'\?', '%s', sql)

    def execute(self, sql: str, params: tuple | list | None = None):
        cur = self.conn.cursor()
        cur.execute(self._convert_sql(sql), params or [])
        return cur

    def executescript(self, script: str):
        if self.kind == 'sqlite':
            return self.conn.executescript(script)
        statements = [s.strip() for s in script.split(';') if s.strip()]
        cur = self.conn.cursor()
        for stmt in statements:
            cur.execute(stmt)
        return cur

    def fetchone(self, sql: str, params: tuple | list | None = None):
        cur = self.execute(sql, params)
        return cur.fetchone()

    def fetchall(self, sql: str, params: tuple | list | None = None):
        cur = self.execute(sql, params)
        return cur.fetchall()



def create_app() -> Flask:
    app = Flask(__name__)
    app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'cambia-esta-clave')
    app.config['DATABASE'] = os.environ.get('DATABASE_URL', str(DB_PATH))
    app.config['PREFERRED_URL_SCHEME'] = 'https'

    login_manager = LoginManager()
    login_manager.login_view = 'login'
    login_manager.login_message = 'Inicia sesión para continuar.'
    login_manager.login_message_category = 'warning'
    login_manager.init_app(app)

    @app.context_processor
    def inject_globals() -> dict[str, Any]:
        return {
            'APP_NAME': APP_NAME,
            'APP_TAGLINE': APP_TAGLINE,
            'SCHOOL_NAME': SCHOOL_NAME,
            'SCHOOL_LOCATION': SCHOOL_LOCATION,
            'now': datetime.now(),
            'formato_monto': formato_monto,
            'estado_cuota': estado_cuota,
            'current_user': current_user,
            'backup_dir': BACKUP_DIR,
            'db_engine': 'PostgreSQL' if is_postgres_url(app.config['DATABASE']) else 'SQLite',
            'current_scope_course': current_course_filter(),
            'selected_admin_course': admin_selected_course(),
            'available_courses': get_available_courses(),
            'predefined_courses': get_available_courses(include_dynamic=True),
            'colegios': get_colegios(),
            'selected_colegio_id': selected_colegio_id(),
            'current_colegio_nombre': current_colegio_nombre(),
            'is_admin_global': current_user.is_authenticated and current_user.is_admin_global(),
            'current_permissions': user_course_permissions(),
        }

    def get_db() -> DBAdapter:
        if 'db' not in g:
            g.db = DBAdapter(app.config['DATABASE'])
        return g.db

    @login_manager.user_loader
    def load_user(user_id: str):
        db = get_db()
        row = db.fetchone('SELECT * FROM usuarios WHERE id = ? AND activo = 1', (int(user_id),))
        return User(row) if row else None

    @app.teardown_appcontext
    def close_db(_exc: Exception | None) -> None:
        db = g.pop('db', None)
        if db is not None:
            db.close()

    app.get_db = get_db  # type: ignore[attr-defined]

    with app.app_context():
        init_db(get_db())
        seed_default_admin(get_db())

    def role_required(*roles: str):
        def decorator(fn):
            @wraps(fn)
            @login_required
            def wrapper(*args, **kwargs):
                if not ((current_user.role in roles) or (('admin_global' in roles or 'admin' in roles) and current_user.is_admin_global())):
                    flash('No tienes permisos para realizar esta acción.', 'danger')
                    return redirect(url_for('dashboard'))
                return fn(*args, **kwargs)
            return wrapper
        return decorator

    def log_audit(accion: str, entidad: str, entidad_id: Any = None, detalle: str = '') -> None:
        """Registra acciones importantes para trazabilidad. No debe romper la operación principal."""
        try:
            db = get_db()
            user_id = int(current_user.id) if current_user.is_authenticated else None
            username = current_user.username if current_user.is_authenticated else 'anonimo'
            colegio_id = selected_colegio_id() if current_user.is_authenticated and current_user.is_admin_global() else getattr(current_user, 'colegio_id', None)
            ip = request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip()[:80]
            db.execute(
                '''INSERT INTO auditoria_acciones
                   (fecha, usuario_id, username, accion, entidad, entidad_id, colegio_id, curso, detalle, ip)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), user_id, username, accion, entidad, str(entidad_id or ''), colegio_id, current_course_filter(), detalle[:500], ip),
            )
            db.commit()
        except Exception:
            try:
                get_db().rollback()
            except Exception:
                pass


    def normalize_course(curso: str | None) -> str:
        return normalize_course_value(curso)


    def colegio_suscripcion_activa(colegio_id: int | None) -> bool:
        if not colegio_id:
            return True
        db = get_db()
        try:
            row = db.fetchone('SELECT estado_suscripcion, fecha_vencimiento FROM colegios WHERE id = ?', (int(colegio_id),))
        except Exception:
            db.rollback()
            return True
        if not row:
            return True
        if (row['estado_suscripcion'] or 'activo') == 'vencido':
            return False
        vencimiento = row['fecha_vencimiento'] if 'fecha_vencimiento' in row.keys() else None
        if vencimiento:
            try:
                return datetime.strptime(vencimiento[:10], '%Y-%m-%d').date() >= datetime.now().date()
            except Exception:
                return True
        return True

    def colegio_id_para_escritura() -> int | None:
        for key in ('colegio_id', 'selected_colegio_id'):
            raw = request.form.get(key) or request.args.get(key)
            if raw and str(raw).isdigit():
                return int(raw)
        if current_user.is_authenticated and not current_user.is_admin_global():
            try:
                return int(current_user.colegio_id) if current_user.colegio_id else None
            except Exception:
                return None
        return selected_colegio_id() if current_user.is_authenticated and current_user.is_admin_global() else None

    @app.before_request
    def bloquear_colegios_vencidos():
        if not current_user.is_authenticated or request.method != 'POST':
            return None
        endpoint = request.endpoint or ''
        if endpoint.startswith(('login', 'logout', 'backups', 'colegios', 'saas_dashboard', 'auditoria', 'usuarios')):
            return None
        colegio_id = colegio_id_para_escritura()
        if colegio_id and not colegio_suscripcion_activa(colegio_id) and not current_user.is_admin_global():
            flash('La suscripción del colegio está vencida. El sistema queda en modo solo lectura hasta regularizar el plan.', 'warning')
            return redirect(request.referrer or url_for('dashboard'))
        return None

    def get_colegios() -> list[Any]:
        db = get_db()
        try:
            return db.fetchall("SELECT id, nombre, ubicacion, activo, plan, estado_suscripcion, fecha_vencimiento FROM colegios WHERE activo = 1 ORDER BY nombre")
        except Exception:
            return []

    def selected_colegio_id() -> int | None:
        raw = request.args.get('colegio_id', '').strip()
        if raw.isdigit():
            return int(raw)
        return None

    def current_colegio_nombre() -> str | None:
        if not current_user.is_authenticated:
            return None
        db = get_db()
        cid = selected_colegio_id() if current_user.is_admin_global() else getattr(current_user, 'colegio_id', None)
        if cid:
            try:
                row = db.fetchone('SELECT nombre FROM colegios WHERE id = ?', (int(cid),))
                if row:
                    return row['nombre']
            except Exception:
                return None
        if current_user.is_admin_global():
            return 'Todos los colegios'
        return None

    def user_course_permissions(user_id: int | None = None) -> list[Any]:
        if not current_user.is_authenticated:
            return []
        uid = user_id or int(current_user.id)
        db = get_db()
        try:
            return db.fetchall("""
                SELECT urc.*, c.nombre AS colegio_nombre
                FROM usuario_roles_curso urc
                INNER JOIN colegios c ON c.id = urc.colegio_id
                WHERE urc.usuario_id = ?
                ORDER BY c.nombre, urc.curso, urc.rol_curso
            """, (uid,))
        except Exception:
            return []

    def user_course_scope() -> str | None:
        if not current_user.is_authenticated or current_user.is_admin_global():
            return None
        curso = request.args.get('curso', '').strip() or (getattr(current_user, 'curso', None) or '').strip()
        return curso or None

    def admin_selected_course() -> str | None:
        if not current_user.is_authenticated or not current_user.is_admin_global():
            return None
        curso = request.args.get('curso', '').strip()
        return curso or None

    def current_course_filter() -> str | None:
        return user_course_scope() or admin_selected_course()

    def get_available_courses(include_dynamic: bool = True) -> list[str]:
        cursos: dict[str, str] = {normalize_course(curso): curso for curso in PREDEFINED_COURSES}
        if include_dynamic and current_user.is_authenticated:
            db = get_db()
            queries = [
                "SELECT DISTINCT curso FROM alumnos WHERE curso IS NOT NULL AND trim(curso) <> ''",
                "SELECT DISTINCT curso FROM actividades WHERE curso IS NOT NULL AND trim(curso) <> ''",
                "SELECT DISTINCT curso FROM movimientos WHERE curso IS NOT NULL AND trim(curso) <> ''",
                "SELECT DISTINCT curso FROM usuario_roles_curso WHERE curso IS NOT NULL AND trim(curso) <> ''",
            ]
            for sql_dyn in queries:
                try:
                    for row in db.fetchall(sql_dyn):
                        valor = (row['curso'] or '').strip()
                        if valor:
                            cursos.setdefault(normalize_course(valor), valor)
                except Exception:
                    continue
        return sorted(cursos.values(), key=lambda x: normalize_course(x))

    def append_scope_filter(sql: str, params: list[Any], curso_expr: str, colegio_expr: str) -> tuple[str, list[Any]]:
        if current_user.is_authenticated and current_user.is_admin_global():
            colegio_id = selected_colegio_id()
            curso = admin_selected_course()
            if colegio_id:
                sql += f" AND {colegio_expr} = ?"
                params.append(colegio_id)
            if curso:
                sql += f" AND lower(trim(COALESCE({curso_expr}, ''))) = lower(trim(?))"
                params.append(curso)
            return sql, params

        permisos = user_course_permissions()
        curso_seleccionado = request.args.get('curso', '').strip()
        colegio_seleccionado = selected_colegio_id()
        if permisos:
            condiciones = []
            for permiso in permisos:
                if colegio_seleccionado and int(permiso['colegio_id']) != colegio_seleccionado:
                    continue
                if curso_seleccionado and normalize_course(permiso['curso']) != normalize_course(curso_seleccionado):
                    continue
                condiciones.append(f"({colegio_expr} = ? AND lower(trim(COALESCE({curso_expr}, ''))) = lower(trim(?)))")
                params.extend([permiso['colegio_id'], permiso['curso']])
            if condiciones:
                sql += ' AND (' + ' OR '.join(condiciones) + ')'
            else:
                sql += ' AND 1=0'
            return sql, params

        colegio_id = getattr(current_user, 'colegio_id', None) or 1
        curso = (getattr(current_user, 'curso', None) or '').strip()
        sql += f" AND {colegio_expr} = ?"
        params.append(colegio_id)
        if curso:
            sql += f" AND lower(trim(COALESCE({curso_expr}, ''))) = lower(trim(?))"
            params.append(curso)
        return sql, params

    def course_filter_sql(sql: str, params: list[Any], alias: str, column: str = 'curso') -> tuple[str, list[Any]]:
        return append_scope_filter(sql, params, f"{alias}.{column}", f"COALESCE({alias}.colegio_id, 1)")

    def movimientos_course_filter_sql(sql: str, params: list[Any], movimiento_alias: str = 'm', actividad_alias: str = 'a', alumno_alias: str = 'al') -> tuple[str, list[Any]]:
        curso_expr = f"COALESCE({movimiento_alias}.curso, {actividad_alias}.curso, {alumno_alias}.curso)"
        colegio_expr = f"COALESCE({movimiento_alias}.colegio_id, {actividad_alias}.colegio_id, {alumno_alias}.colegio_id, 1)"
        return append_scope_filter(sql, params, curso_expr, colegio_expr)

    def ensure_course_access(curso: str | None, colegio_id: int | None = None) -> bool:
        if not current_user.is_authenticated or current_user.is_admin_global():
            return True
        colegio_id = colegio_id or getattr(current_user, 'colegio_id', None) or 1
        permisos = user_course_permissions()
        if permisos:
            return any(int(p['colegio_id']) == int(colegio_id) and normalize_course(p['curso']) == normalize_course(curso) for p in permisos)
        scope = (getattr(current_user, 'curso', None) or '').strip()
        return (not scope or normalize_course(scope) == normalize_course(curso)) and int(colegio_id) == int(getattr(current_user, 'colegio_id', None) or 1)

    def mes_esta_cerrado(db: DBAdapter, colegio_id: int | None, curso: str | None, fecha_o_mes: str | None) -> bool:
        """Retorna True si existe un cierre mensual para colegio+curso+mes."""
        if not colegio_id or not curso or not fecha_o_mes:
            return False
        mes = (fecha_o_mes or '')[:7]
        if not re.match(r'^\d{4}-\d{2}$', mes):
            return False
        row = db.fetchone(
            'SELECT id FROM cierres_mensuales WHERE colegio_id = ? AND lower(trim(curso)) = lower(trim(?)) AND mes = ? LIMIT 1',
            (int(colegio_id), curso, mes),
        )
        return bool(row)

    def assert_mes_abierto(db: DBAdapter, colegio_id: int | None, curso: str | None, fecha_o_mes: str | None) -> None:
        if mes_esta_cerrado(db, colegio_id, curso, fecha_o_mes) and not current_user.is_admin_global():
            raise ValueError('El mes ya está cerrado para este colegio y curso. Solo admin puede modificarlo.')

    def cuota_configurada(db: DBAdapter, colegio_id: int, curso: str, mes: str) -> float:
        row = db.fetchone(
            'SELECT monto FROM cuotas_mensuales WHERE colegio_id = ? AND lower(trim(curso)) = lower(trim(?)) AND mes = ?',
            (colegio_id, curso, mes),
        )
        return float(row['monto']) if row else 0.0

    def resolve_colegio_for_course(curso: str | None) -> int:
        if current_user.is_authenticated and current_user.is_admin_global():
            raw = (request.form.get('colegio_id') or request.args.get('colegio_id') or '').strip()
            return int(raw) if raw.isdigit() else 1
        permisos = user_course_permissions()
        for p in permisos:
            if normalize_course(p['curso']) == normalize_course(curso):
                return int(p['colegio_id'])
        return int(getattr(current_user, 'colegio_id', None) or 1)

    def fetch_pago_permitido(db: DBAdapter, pago_id: int):
        pago = db.fetchone(
            """
            SELECT p.*, a.curso AS curso_alumno, a.colegio_id AS colegio_id
            FROM pagos_alumnos p
            INNER JOIN alumnos a ON a.id = p.alumno_id
            WHERE p.id = ?
            """,
            (pago_id,),
        )
        if pago and ensure_course_access(pago['curso_alumno'], pago['colegio_id'] if 'colegio_id' in pago.keys() else 1):
            return pago
        return None

    def fetch_movimiento_permitido(db: DBAdapter, movimiento_id: int):
        movimiento = db.fetchone(
            """
            SELECT m.*, COALESCE(m.curso, al.curso, ac.curso) AS curso_ref, COALESCE(m.colegio_id, al.colegio_id, ac.colegio_id, 1) AS colegio_ref
            FROM movimientos m
            LEFT JOIN alumnos al ON al.id = m.alumno_id
            LEFT JOIN actividades ac ON ac.id = m.actividad_id
            WHERE m.id = ?
            """,
            (movimiento_id,),
        )
        if movimiento and ensure_course_access(movimiento['curso_ref'], movimiento['colegio_ref'] if 'colegio_ref' in movimiento.keys() else 1):
            return movimiento
        return None

    def fetch_actividad_permitida(db: DBAdapter, actividad_id: int):
        actividad = db.fetchone('SELECT * FROM actividades WHERE id = ?', (actividad_id,))
        if actividad and ensure_course_access(actividad['curso'], actividad['colegio_id'] if 'colegio_id' in actividad.keys() else 1):
            return actividad
        return None

    def fetch_alumno_permitido(db: DBAdapter, alumno_id: int):
        alumno = db.fetchone(
            """
            SELECT a.*, c.nombre AS colegio_nombre
            FROM alumnos a
            LEFT JOIN colegios c ON c.id = a.colegio_id
            WHERE a.id = ?
            """,
            (alumno_id,),
        )
        if alumno and ensure_course_access(alumno['curso'], alumno['colegio_id'] if 'colegio_id' in alumno.keys() else 1):
            return alumno
        return None

    def resolver_curso_operacion(db: DBAdapter, alumno_id: int | None = None, actividad_id: int | None = None, curso_form: str | None = None) -> str | None:
        curso = (curso_form or '').strip() or user_course_scope()
        if alumno_id:
            alumno = fetch_alumno_permitido(db, alumno_id)
            if not alumno:
                raise ValueError('Alumno no encontrado para tu curso.')
            curso = alumno['curso']
        if actividad_id:
            actividad = fetch_actividad_permitida(db, actividad_id)
            if not actividad:
                raise ValueError('Actividad no encontrada para tu curso.')
            if curso and normalize_course(curso) != normalize_course(actividad['curso']):
                raise ValueError('El curso del movimiento no coincide con el de la actividad.')
            curso = actividad['curso']
        if not ensure_course_access(curso, resolve_colegio_for_course(curso)):
            raise ValueError('No puedes registrar información en otro curso.')
        return curso

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for('dashboard'))
        if request.method == 'POST':
            username = request.form.get('username', '').strip().lower()
            email = request.form.get('email', '').strip().lower()
            password = request.form.get('password', '')
            db = get_db()
            row = db.fetchone("SELECT * FROM usuarios WHERE (lower(username) = ? OR lower(COALESCE(email,'')) = ?) AND activo = 1", (username, username))
            if row and check_password_hash(row['password_hash'], password):
                login_user(User(row), remember=True)
                log_audit('login', 'sesion', row['id'], 'Inicio de sesión correcto')
                flash(f'Bienvenido, {row["nombre"]}.', 'success')
                next_url = request.args.get('next')
                if next_url and urlparse(next_url).netloc == '':
                    return redirect(next_url)
                return redirect(url_for('dashboard'))
            flash('Usuario/correo o contraseña incorrectos.', 'danger')
        return render_template('login.html')

    @app.route('/recuperar-contrasena', methods=['GET', 'POST'])
    def recuperar_contrasena():
        if current_user.is_authenticated:
            return redirect(url_for('dashboard'))
        token_generado = None
        if request.method == 'POST':
            identificador = request.form.get('identificador', '').strip().lower()
            db = get_db()
            user = db.fetchone("SELECT id, username, email, nombre FROM usuarios WHERE lower(username)=? OR lower(COALESCE(email,''))=?", (identificador, identificador))
            if user:
                token_generado = secrets.token_urlsafe(32)
                expira = (datetime.now() + timedelta(hours=2)).strftime('%Y-%m-%d %H:%M:%S')
                db.execute('DELETE FROM password_reset_tokens WHERE usuario_id=?', (user['id'],))
                db.execute('INSERT INTO password_reset_tokens (usuario_id, token, expira_en, usado) VALUES (?, ?, ?, 0)', (user['id'], token_generado, expira))
                db.commit()
                log_audit('crear', 'password_reset', user['id'], 'Token de recuperación generado')
                flash('Token generado. Cópialo y úsalo para restablecer la contraseña.', 'success')
            else:
                flash('Si el usuario/correo existe, se podrá generar un token de recuperación.', 'info')
        return render_template('recuperar_contrasena.html', token=token_generado)

    @app.route('/restablecer-contrasena/<token>', methods=['GET', 'POST'])
    def restablecer_contrasena(token: str):
        if current_user.is_authenticated:
            return redirect(url_for('dashboard'))
        db = get_db()
        row = db.fetchone("SELECT t.id, t.usuario_id, t.expira_en, t.usado, u.username, u.nombre FROM password_reset_tokens t INNER JOIN usuarios u ON u.id = t.usuario_id WHERE t.token = ?", (token,))
        valido = False
        if row and not row['usado']:
            try:
                valido = datetime.strptime(row['expira_en'], '%Y-%m-%d %H:%M:%S') >= datetime.now()
            except Exception:
                valido = False
        if not valido:
            flash('El enlace de recuperación no existe, expiró o ya fue utilizado.', 'danger')
            return redirect(url_for('login'))
        if request.method == 'POST':
            password = request.form.get('password', '')
            confirmacion = request.form.get('confirmacion', '')
            if len(password) < 8:
                flash('La nueva contraseña debe tener al menos 8 caracteres.', 'danger')
            elif password != confirmacion:
                flash('Las contraseñas no coinciden.', 'danger')
            else:
                db.execute('UPDATE usuarios SET password_hash=? WHERE id=?', (generate_password_hash(password), row['usuario_id']))
                db.execute('UPDATE password_reset_tokens SET usado=1 WHERE id=?', (row['id'],))
                db.commit()
                log_audit('actualizar', 'usuario', row['usuario_id'], 'Contraseña restablecida con token')
                flash('Contraseña actualizada. Ya puedes iniciar sesión.', 'success')
                return redirect(url_for('login'))
        return render_template('restablecer_contrasena.html', token=token, usuario=row)

    @app.get('/logout')
    @login_required
    def logout():
        log_audit('logout', 'sesion', current_user.id, 'Cierre de sesión')
        logout_user()
        flash('Sesión cerrada.', 'success')
        return redirect(url_for('login'))

    @app.get('/healthz')
    def healthz():
        return {'status': 'ok'}, 200

    @app.route('/')
    def index():
        return redirect(url_for('dashboard' if current_user.is_authenticated else 'login'))

    @app.route('/dashboard')
    @login_required
    def dashboard():
        db = get_db()
        resumen_sql = """
            SELECT
                COALESCE(SUM(CASE WHEN tipo='ingreso' THEN monto ELSE 0 END),0) ingresos,
                COALESCE(SUM(CASE WHEN tipo='gasto' THEN monto ELSE 0 END),0) gastos,
                COUNT(*) cantidad
            FROM movimientos m
            LEFT JOIN actividades a ON a.id = m.actividad_id
            LEFT JOIN alumnos al ON al.id = m.alumno_id
            WHERE 1=1
        """
        resumen_params: list[Any] = []
        resumen_sql, resumen_params = movimientos_course_filter_sql(resumen_sql, resumen_params)
        resumen = db.fetchone(resumen_sql, resumen_params)

        reporte_sql = """
            SELECT substr(m.fecha, 1, 7) AS mes,
                   COALESCE(SUM(CASE WHEN m.tipo = 'ingreso' THEN m.monto ELSE 0 END), 0) AS ingresos,
                   COALESCE(SUM(CASE WHEN m.tipo = 'gasto' THEN m.monto ELSE 0 END), 0) AS gastos
            FROM movimientos m
            LEFT JOIN actividades a ON a.id = m.actividad_id
            LEFT JOIN alumnos al ON al.id = m.alumno_id
            WHERE 1=1
        """
        reporte_params: list[Any] = []
        reporte_sql, reporte_params = movimientos_course_filter_sql(reporte_sql, reporte_params)
        reporte_sql += ' GROUP BY substr(m.fecha, 1, 7) ORDER BY mes ASC'
        reporte = db.fetchall(reporte_sql, reporte_params)
        mes = request.args.get('mes') or datetime.today().strftime('%Y-%m')
        alertas = obtener_alertas_morosidad(db, mes, current_course_filter(), selected_colegio_id() if current_user.is_admin_global() else getattr(current_user, 'colegio_id', None))
        ultimos_sql = """
            SELECT m.id, m.fecha, m.tipo, m.concepto, m.monto, COALESCE(a.nombre, '-') AS actividad
            FROM movimientos m
            LEFT JOIN actividades a ON a.id = m.actividad_id
            LEFT JOIN alumnos al ON al.id = m.alumno_id
            WHERE 1=1
        """
        ultimos_params: list[Any] = []
        ultimos_sql, ultimos_params = movimientos_course_filter_sql(ultimos_sql, ultimos_params)
        ultimos_sql += ' ORDER BY m.fecha DESC, m.id DESC LIMIT 8'
        ultimos = db.fetchall(ultimos_sql, ultimos_params)
        backups = listar_backups()[:5]

        resumen_mes_sql = """
            SELECT
                COALESCE(SUM(CASE WHEN m.tipo='ingreso' THEN m.monto ELSE 0 END),0) ingresos_mes,
                COALESCE(SUM(CASE WHEN m.tipo='gasto' THEN m.monto ELSE 0 END),0) gastos_mes,
                COUNT(*) movimientos_mes
            FROM movimientos m
            LEFT JOIN actividades a ON a.id = m.actividad_id
            LEFT JOIN alumnos al ON al.id = m.alumno_id
            WHERE substr(m.fecha, 1, 7) = ?
        """
        resumen_mes_params: list[Any] = [mes]
        resumen_mes_sql, resumen_mes_params = movimientos_course_filter_sql(resumen_mes_sql, resumen_mes_params)
        resumen_mes = db.fetchone(resumen_mes_sql, resumen_mes_params)
        alumnos_sql = 'SELECT COUNT(*) AS total FROM alumnos a WHERE activo = 1'
        alumnos_params: list[Any] = []
        alumnos_sql, alumnos_params = course_filter_sql(alumnos_sql, alumnos_params, 'a')
        alumnos_activos = db.fetchone(alumnos_sql, alumnos_params)
        cuotas = resumen_cuotas_por_alumno(db, mes, current_course_filter(), selected_colegio_id() if current_user.is_admin_global() else getattr(current_user, 'colegio_id', None))
        total_esperado = sum(float(f['cuota_mensual']) for f in cuotas if f['activo'])
        total_pagado = sum(float(f['pagado']) for f in cuotas if f['activo'])
        deuda_total = sum(max(float(f['cuota_mensual']) - float(f['pagado']), 0) for f in cuotas if f['activo'])
        alumnos_pagados = 0
        alumnos_parciales = 0
        alumnos_deuda = 0
        for fila in cuotas:
            if not fila['activo']:
                continue
            estado, _icono = estado_cuota(fila['cuota_mensual'], fila['pagado'])
            if estado == 'Pagado':
                alumnos_pagados += 1
            elif estado == 'Parcial':
                alumnos_parciales += 1
            else:
                alumnos_deuda += 1

        ingresos_mes = float(resumen_mes['ingresos_mes'] or 0)
        gastos_mes = float(resumen_mes['gastos_mes'] or 0)
        balance_total = float(resumen['ingresos'] or 0) - float(resumen['gastos'] or 0)
        balance_mes = ingresos_mes - gastos_mes
        cumplimiento = round((total_pagado / total_esperado) * 100, 1) if total_esperado else 100.0
        ultimo_mes = dict(reporte[-1]) if reporte else {'mes': mes, 'ingresos': 0, 'gastos': 0}
        quick_actions = [
            {'label': 'Nuevo pago', 'href': url_for('pagos_new'), 'icon': '💳', 'hint': 'Registrar cuota o aporte'},
            {'label': 'Nuevo ingreso', 'href': url_for('movimientos_new') + '?tipo=ingreso', 'icon': '➕', 'hint': 'Agregar ingreso manual'},
            {'label': 'Nuevo gasto', 'href': url_for('movimientos_new') + '?tipo=gasto', 'icon': '🧾', 'hint': 'Registrar egreso'},
            {'label': 'Ver cuotas', 'href': url_for('cuotas_view', mes=mes), 'icon': '📌', 'hint': 'Revisar estado mensual'},
        ]

        dashboard_stats = {
            'balance_total': balance_total,
            'balance_mes': balance_mes,
            'deuda_total': deuda_total,
            'alumnos_activos': int(alumnos_activos['total'] or 0),
            'alumnos_pagados': alumnos_pagados,
            'alumnos_parciales': alumnos_parciales,
            'alumnos_deuda': alumnos_deuda,
            'ingresos_mes': ingresos_mes,
            'gastos_mes': gastos_mes,
            'movimientos_mes': int(resumen_mes['movimientos_mes'] or 0),
            'cumplimiento': cumplimiento,
            'ultimo_mes': ultimo_mes,
            'total_esperado': total_esperado,
            'total_pagado': total_pagado,
        }

        resumen_colegios = []
        if current_user.is_admin_global():
            # IMPORTANTE: no unir alumnos y movimientos en la misma consulta agregada.
            # Si se hace, cada movimiento se multiplica por la cantidad de alumnos del colegio
            # y el resumen muestra ingresos/gastos inflados. Se agregan en subconsultas separadas.
            resumen_colegios = db.fetchall(
                """
                SELECT c.id, c.nombre,
                       COALESCE(al.alumnos, 0) AS alumnos,
                       COALESCE(mv.ingresos, 0) AS ingresos,
                       COALESCE(mv.gastos, 0) AS gastos
                FROM colegios c
                LEFT JOIN (
                    SELECT COALESCE(colegio_id, 1) AS colegio_id, COUNT(*) AS alumnos
                    FROM alumnos
                    WHERE activo = 1
                    GROUP BY COALESCE(colegio_id, 1)
                ) al ON al.colegio_id = c.id
                LEFT JOIN (
                    SELECT COALESCE(colegio_id, 1) AS colegio_id,
                           COALESCE(SUM(CASE WHEN tipo='ingreso' THEN monto ELSE 0 END), 0) AS ingresos,
                           COALESCE(SUM(CASE WHEN tipo='gasto' THEN monto ELSE 0 END), 0) AS gastos
                    FROM movimientos
                    GROUP BY COALESCE(colegio_id, 1)
                ) mv ON mv.colegio_id = c.id
                WHERE c.activo = 1
                ORDER BY c.nombre
                """
            )

        return render_template(
            'dashboard.html',
            resumen=resumen,
            reporte=reporte,
            mes=mes,
            alertas=alertas,
            ultimos=ultimos,
            backups=backups,
            dashboard_stats=dashboard_stats,
            quick_actions=quick_actions,
            resumen_colegios=resumen_colegios,
        )


    @app.route('/buscar')
    @login_required
    def buscar_global():
        db = get_db()
        q = request.args.get('q', '').strip()
        resultados = {'alumnos': [], 'pagos': [], 'movimientos': [], 'actividades': []}
        if q:
            like = sql_like_ci(q)
            alumnos_sql = """
                SELECT a.id, a.nombre, a.curso, c.nombre AS colegio_nombre, a.apoderado, a.telefono
                FROM alumnos a
                LEFT JOIN colegios c ON c.id = COALESCE(a.colegio_id, 1)
                WHERE (LOWER(COALESCE(a.nombre,'')) LIKE ? OR LOWER(COALESCE(a.apoderado,'')) LIKE ? OR LOWER(COALESCE(a.telefono,'')) LIKE ? OR LOWER(COALESCE(a.curso,'')) LIKE ?)
            """
            alumnos_params = [like, like, like, like]
            alumnos_sql, alumnos_params = course_filter_sql(alumnos_sql, alumnos_params, 'a')
            alumnos_sql += ' ORDER BY a.nombre LIMIT 25'
            resultados['alumnos'] = db.fetchall(alumnos_sql, alumnos_params)

            movimientos_sql = """
                SELECT m.id, m.fecha, m.tipo, m.concepto, m.monto, COALESCE(m.curso, al.curso, ac.curso) AS curso, c.nombre AS colegio_nombre
                FROM movimientos m
                LEFT JOIN alumnos al ON al.id = m.alumno_id
                LEFT JOIN actividades ac ON ac.id = m.actividad_id
                LEFT JOIN colegios c ON c.id = COALESCE(m.colegio_id, al.colegio_id, ac.colegio_id, 1)
                WHERE (LOWER(COALESCE(m.concepto,'')) LIKE ? OR LOWER(COALESCE(m.observacion,'')) LIKE ? OR LOWER(COALESCE(al.nombre,'')) LIKE ? OR LOWER(COALESCE(ac.nombre,'')) LIKE ?)
            """
            movimientos_params = [like, like, like, like]
            movimientos_sql, movimientos_params = movimientos_course_filter_sql(movimientos_sql, movimientos_params)
            movimientos_sql += ' ORDER BY m.fecha DESC, m.id DESC LIMIT 25'
            resultados['movimientos'] = db.fetchall(movimientos_sql, movimientos_params)

            pagos_sql = """
                SELECT p.id, p.fecha, p.mes, p.monto, a.nombre AS alumno, a.curso, c.nombre AS colegio_nombre
                FROM pagos_alumnos p
                INNER JOIN alumnos a ON a.id = p.alumno_id
                LEFT JOIN colegios c ON c.id = COALESCE(a.colegio_id, 1)
                WHERE (LOWER(COALESCE(a.nombre,'')) LIKE ? OR LOWER(COALESCE(p.observacion,'')) LIKE ? OR p.mes LIKE ?)
            """
            pagos_params = [like, like, f'%{q}%']
            pagos_sql, pagos_params = course_filter_sql(pagos_sql, pagos_params, 'a')
            pagos_sql += ' ORDER BY p.fecha DESC, p.id DESC LIMIT 25'
            resultados['pagos'] = db.fetchall(pagos_sql, pagos_params)

            actividades_sql = """
                SELECT a.id, a.nombre, a.fecha, a.curso, c.nombre AS colegio_nombre, a.estado
                FROM actividades a
                LEFT JOIN colegios c ON c.id = COALESCE(a.colegio_id, 1)
                WHERE (LOWER(COALESCE(a.nombre,'')) LIKE ? OR LOWER(COALESCE(a.descripcion,'')) LIKE ? OR LOWER(COALESCE(a.curso,'')) LIKE ?)
            """
            actividades_params = [like, like, like]
            actividades_sql, actividades_params = course_filter_sql(actividades_sql, actividades_params, 'a')
            actividades_sql += ' ORDER BY a.fecha DESC, a.id DESC LIMIT 25'
            resultados['actividades'] = db.fetchall(actividades_sql, actividades_params)

        return render_template('buscar.html', q=q, resultados=resultados)

    @app.route('/movimientos/export/<fmt>')
    @login_required
    def movimientos_export(fmt: str):
        db = get_db()
        tipo = request.args.get('tipo', 'Todos')
        mes = request.args.get('mes', '')
        q = request.args.get('q', '').strip()
        movimientos = obtener_movimientos_filtrados(db, tipo=tipo, mes=mes, q=q, curso_scope=current_course_filter(), colegio_scope=selected_colegio_id() if current_user.is_admin_global() else getattr(current_user, 'colegio_id', None))
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre = f'movimientos_{ts}'
        if fmt == 'csv':
            sio = StringIO()
            writer = csv.writer(sio)
            writer.writerow(['ID', 'Fecha', 'Tipo', 'Concepto', 'Actividad', 'Alumno', 'Origen', 'Monto', 'Observación'])
            for row in movimientos:
                writer.writerow([row['id'], row['fecha'], row['tipo'], row['concepto'], row['actividad'], row.get('alumno', '-'), row['origen'], row['monto'], row['observacion']])
            data = BytesIO(sio.getvalue().encode('utf-8-sig'))
            return send_file(data, mimetype='text/csv', as_attachment=True, download_name=f'{nombre}.csv')
        if fmt == 'xlsx':
            wb = Workbook()
            ws = wb.active
            ws.title = 'Movimientos'
            ws.append(['ID', 'Fecha', 'Tipo', 'Concepto', 'Actividad', 'Alumno', 'Origen', 'Monto', 'Observación'])
            for row in movimientos:
                ws.append([row['id'], row['fecha'], row['tipo'], row['concepto'], row['actividad'], row.get('alumno', '-'), row['origen'], float(row['monto']), row['observacion']])
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)
            for column in ['A','B','C','D','E','F','G','H']:
                ws.column_dimensions[column].width = 18 if column != 'D' else 34
            data = BytesIO()
            wb.save(data)
            data.seek(0)
            return send_file(data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'{nombre}.xlsx')
        if fmt == 'pdf':
            data = exportar_movimientos_pdf(movimientos, SCHOOL_NAME, SCHOOL_LOCATION, {'Tipo': tipo, 'Mes': mes or 'Todos', 'Búsqueda': q or 'Todos'})
            return send_file(data, mimetype='application/pdf', as_attachment=True, download_name=f'{nombre}.pdf')
        flash('Formato de exportación no soportado.', 'danger')
        return redirect(url_for('movimientos_list', tipo=tipo, mes=mes, q=q))

    @app.route('/backups')
    @role_required('admin')
    def backups_list():
        backups = listar_backups()
        return render_template('backups.html', backups=backups)

    @app.post('/backups/crear')
    @role_required('admin')
    def backups_create():
        try:
            ruta = crear_backup_db(app.config['DATABASE'])
            log_audit('crear_respaldo', 'backup', ruta.name, 'Respaldo manual creado desde el panel')
            flash(f'Respaldo creado: {ruta.name}', 'success')
        except Exception as exc:
            flash(f'No se pudo crear el respaldo: {exc}', 'danger')
        return redirect(request.referrer or url_for('backups_list'))

    @app.route('/admin/plataforma')
    @role_required('admin')
    def saas_dashboard():
        db = get_db()
        resumen = db.fetchall("""
            SELECT c.id, c.nombre, c.plan, c.estado_suscripcion, c.fecha_vencimiento, c.activo,
                   COALESCE(al.alumnos, 0) AS alumnos,
                   COALESCE(us.usuarios, 0) AS usuarios,
                   COALESCE(pg.pagos, 0) AS pagos,
                   COALESCE(pg.monto_pagos, 0) AS monto_pagos
            FROM colegios c
            LEFT JOIN (
                SELECT COALESCE(colegio_id, 1) AS colegio_id, COUNT(*) AS alumnos
                FROM alumnos
                GROUP BY COALESCE(colegio_id, 1)
            ) al ON al.colegio_id = c.id
            LEFT JOIN (
                SELECT COALESCE(colegio_id, 1) AS colegio_id, COUNT(*) AS usuarios
                FROM usuarios
                WHERE role <> 'admin'
                GROUP BY COALESCE(colegio_id, 1)
            ) us ON us.colegio_id = c.id
            LEFT JOIN (
                SELECT COALESCE(a.colegio_id, 1) AS colegio_id, COUNT(p.id) AS pagos, COALESCE(SUM(p.monto), 0) AS monto_pagos
                FROM pagos_alumnos p
                INNER JOIN alumnos a ON a.id = p.alumno_id
                GROUP BY COALESCE(a.colegio_id, 1)
            ) pg ON pg.colegio_id = c.id
            GROUP BY c.id, c.nombre, c.plan, c.estado_suscripcion, c.fecha_vencimiento, c.activo, al.alumnos, us.usuarios, pg.pagos, pg.monto_pagos
            ORDER BY c.nombre
        """)
        totales = {
            'colegios': len(resumen),
            'activos': sum(1 for r in resumen if r['activo'] and (r['estado_suscripcion'] or 'activo') == 'activo'),
            'vencidos': sum(1 for r in resumen if (r['estado_suscripcion'] or 'activo') == 'vencido'),
            'alumnos': sum(int(r['alumnos'] or 0) for r in resumen),
            'usuarios': sum(int(r['usuarios'] or 0) for r in resumen),
            'pagos': sum(int(r['pagos'] or 0) for r in resumen),
        }
        return render_template('saas_dashboard.html', resumen=resumen, totales=totales)

    @app.route('/colegios')
    @role_required('admin')
    def colegios_list():
        db = get_db()
        colegios_rows = db.fetchall('SELECT id, nombre, ubicacion, activo, plan, estado_suscripcion, fecha_vencimiento FROM colegios ORDER BY nombre')
        return render_template('colegios_list.html', colegios=colegios_rows)

    @app.route('/colegios/nuevo', methods=['GET', 'POST'])
    @role_required('admin')
    def colegios_new():
        db = get_db()
        if request.method == 'POST':
            nombre = request.form.get('nombre', '').strip()
            ubicacion = request.form.get('ubicacion', '').strip()
            activo = 1 if request.form.get('activo') == 'on' else 0
            plan = request.form.get('plan', 'basico').strip().lower()
            estado_suscripcion = request.form.get('estado_suscripcion', 'activo').strip().lower()
            fecha_vencimiento = request.form.get('fecha_vencimiento', '').strip() or None
            if plan not in PLANES_COLEGIO:
                plan = 'basico'
            if estado_suscripcion not in ESTADOS_SUSCRIPCION:
                estado_suscripcion = 'activo'
            if not nombre:
                flash('El nombre del colegio es obligatorio.', 'danger')
            else:
                db.execute('INSERT INTO colegios (nombre, ubicacion, activo, plan, estado_suscripcion, fecha_vencimiento) VALUES (?, ?, ?, ?, ?, ?)', (nombre, ubicacion, activo, plan, estado_suscripcion, fecha_vencimiento))
                db.commit()
                log_audit('crear', 'colegio', db.fetchone('SELECT MAX(id) AS id FROM colegios')['id'], nombre)
                flash('Colegio creado.', 'success')
                return redirect(url_for('colegios_list'))
        return render_template('colegios_form.html', colegio=None)

    @app.route('/colegios/<int:colegio_id>/editar', methods=['GET', 'POST'])
    @role_required('admin')
    def colegios_edit(colegio_id: int):
        db = get_db()
        colegio = db.fetchone('SELECT id, nombre, ubicacion, activo, plan, estado_suscripcion, fecha_vencimiento FROM colegios WHERE id=?', (colegio_id,))
        if not colegio:
            flash('Colegio no encontrado.', 'danger')
            return redirect(url_for('colegios_list'))
        if request.method == 'POST':
            nombre = request.form.get('nombre', '').strip()
            ubicacion = request.form.get('ubicacion', '').strip()
            activo = 1 if request.form.get('activo') == 'on' else 0
            plan = request.form.get('plan', colegio['plan'] if 'plan' in colegio.keys() else 'basico').strip().lower()
            estado_suscripcion = request.form.get('estado_suscripcion', colegio['estado_suscripcion'] if 'estado_suscripcion' in colegio.keys() else 'activo').strip().lower()
            fecha_vencimiento = request.form.get('fecha_vencimiento', '').strip() or None
            if plan not in PLANES_COLEGIO:
                plan = 'basico'
            if estado_suscripcion not in ESTADOS_SUSCRIPCION:
                estado_suscripcion = 'activo'
            if not nombre:
                flash('El nombre del colegio es obligatorio.', 'danger')
            else:
                db.execute('UPDATE colegios SET nombre=?, ubicacion=?, activo=?, plan=?, estado_suscripcion=?, fecha_vencimiento=? WHERE id=?', (nombre, ubicacion, activo, plan, estado_suscripcion, fecha_vencimiento, colegio_id))
                db.commit()
                log_audit('editar', 'colegio', colegio_id, nombre)
                flash('Colegio actualizado.', 'success')
                return redirect(url_for('colegios_list'))
        return render_template('colegios_form.html', colegio=colegio)

    @app.get('/backups/<path:nombre>')
    @role_required('admin')
    def backups_download(nombre: str):
        ruta = BACKUP_DIR / nombre
        if not ruta.exists() or ruta.parent != BACKUP_DIR:
            flash('Respaldo no encontrado.', 'danger')
            return redirect(url_for('backups_list'))
        return send_file(ruta, as_attachment=True, download_name=ruta.name)

    @app.post('/backups/<path:nombre>/restaurar')
    @role_required('admin')
    def backups_restore(nombre: str):
        if is_postgres_url(app.config['DATABASE']):
            flash('La restauración automática desde el panel solo está disponible para SQLite.', 'warning')
            return redirect(url_for('backups_list'))
        safe_name = secure_filename(nombre)
        if safe_name != nombre:
            flash('Nombre de respaldo inválido.', 'danger')
            return redirect(url_for('backups_list'))
        origen = BACKUP_DIR / safe_name
        destino = Path(app.config['DATABASE'])
        if not origen.exists() or ruta_fuera_de_backups(origen) or origen.suffix != '.db':
            flash('Solo se pueden restaurar respaldos .db válidos.', 'danger')
            return redirect(url_for('backups_list'))
        try:
            respaldo_previo = crear_backup_db(destino)
            db_abierta = g.pop('db', None)
            if db_abierta is not None:
                db_abierta.close()
            shutil.copy2(origen, destino)
            flash(f'Respaldo restaurado: {safe_name}. Se creó backup previo: {respaldo_previo.name}', 'success')
        except Exception as exc:
            flash(f'No se pudo restaurar el respaldo: {exc}', 'danger')
        return redirect(url_for('backups_list'))

    @app.post('/backups/<path:nombre>/eliminar')
    @role_required('admin')
    def backups_delete(nombre: str):
        safe_name = secure_filename(nombre)
        if safe_name != nombre:
            flash('Nombre de respaldo inválido.', 'danger')
            return redirect(url_for('backups_list'))
        ruta = BACKUP_DIR / safe_name
        if not ruta.exists() or ruta_fuera_de_backups(ruta):
            flash('Respaldo no encontrado.', 'danger')
            return redirect(url_for('backups_list'))
        try:
            ruta.unlink()
            log_audit('eliminar_respaldo', 'backup', safe_name, 'Respaldo eliminado desde el panel')
            flash(f'Respaldo eliminado: {safe_name}', 'success')
        except Exception as exc:
            flash(f'No se pudo eliminar el respaldo: {exc}', 'danger')
        return redirect(url_for('backups_list'))


    def calcular_resumen_cierre(db: DBAdapter, colegio_id: int, curso: str, mes: str) -> dict[str, Any]:
        resumen_mov = db.fetchone(
            """
            SELECT
                COALESCE(SUM(CASE WHEN tipo='ingreso' THEN monto ELSE 0 END),0) AS ingresos,
                COALESCE(SUM(CASE WHEN tipo='gasto' THEN monto ELSE 0 END),0) AS gastos
            FROM movimientos
            WHERE COALESCE(colegio_id, 1) = ?
              AND lower(trim(COALESCE(curso, ''))) = lower(trim(?))
              AND substr(fecha, 1, 7) = ?
            """,
            (colegio_id, curso, mes),
        )
        cuotas = resumen_cuotas_por_alumno(db, mes, curso, colegio_id)
        alumnos_activos = sum(1 for f in cuotas if f['activo'])
        total_pagado = sum(float(f['pagado'] or 0) for f in cuotas if f['activo'])
        deuda_total = sum(max(float(f['cuota_mensual'] or 0) - float(f['pagado'] or 0), 0) for f in cuotas if f['activo'])
        ingresos = float(resumen_mov['ingresos'] or 0)
        gastos = float(resumen_mov['gastos'] or 0)
        return {'ingresos': ingresos, 'gastos': gastos, 'saldo': ingresos - gastos, 'alumnos_activos': alumnos_activos, 'total_pagado': total_pagado, 'deuda_total': deuda_total}

    @app.route('/cierres')
    @login_required
    def cierres_list():
        db = get_db()
        mes = request.args.get('mes') or datetime.today().strftime('%Y-%m')
        colegio_scope = selected_colegio_id() if current_user.is_admin_global() else getattr(current_user, 'colegio_id', None)
        curso_scope = current_course_filter()
        sql = """
            SELECT cm.*, c.nombre AS colegio_nombre
            FROM cierres_mensuales cm
            LEFT JOIN colegios c ON c.id = cm.colegio_id
            WHERE 1=1
        """
        params: list[Any] = []
        if mes:
            sql += ' AND cm.mes = ?'
            params.append(mes)
        if colegio_scope:
            sql += ' AND cm.colegio_id = ?'
            params.append(colegio_scope)
        if curso_scope:
            sql += " AND lower(trim(cm.curso)) = lower(trim(?))"
            params.append(curso_scope)
        sql += ' ORDER BY cm.mes DESC, colegio_nombre, cm.curso'
        cierres = db.fetchall(sql, params)
        return render_template('cierres.html', cierres=cierres, mes=mes)

    @app.post('/cierres/crear')
    @login_required
    def cierres_create():
        if not current_user.can_edit():
            flash('No tienes permisos para crear cierres.', 'danger')
            return redirect(url_for('cierres_list'))
        db = get_db()
        mes = request.form.get('mes') or datetime.today().strftime('%Y-%m')
        curso = (request.form.get('curso') or current_course_filter() or '').strip()
        colegio_raw = request.form.get('colegio_id') or (selected_colegio_id() if current_user.is_admin_global() else getattr(current_user, 'colegio_id', None))
        try:
            colegio_id = int(colegio_raw)
        except Exception:
            colegio_id = 0
        if not colegio_id or not curso:
            flash('Debes seleccionar colegio y curso para cerrar el mes.', 'danger')
            return redirect(url_for('cierres_list', mes=mes))
        if not current_user.is_admin_global() and int(getattr(current_user, 'colegio_id', 0) or 0) != colegio_id:
            flash('No tienes permiso para cerrar este colegio.', 'danger')
            return redirect(url_for('cierres_list', mes=mes))
        if not ensure_course_access(curso, colegio_id):
            flash('No tienes permiso para cerrar este curso.', 'danger')
            return redirect(url_for('cierres_list', mes=mes))
        resumen = calcular_resumen_cierre(db, colegio_id, curso, mes)
        observacion = (request.form.get('observacion') or '').strip()
        try:
            db.execute(
                """
                INSERT INTO cierres_mensuales
                (colegio_id, curso, mes, ingresos, gastos, saldo, alumnos_activos, total_pagado, deuda_total, creado_por, creado_en, observacion)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (colegio_id, curso, mes, resumen['ingresos'], resumen['gastos'], resumen['saldo'], resumen['alumnos_activos'], resumen['total_pagado'], resumen['deuda_total'], current_user.username, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), observacion),
            )
            db.commit()
            log_audit('crear', 'cierre_mensual', f'{colegio_id}-{curso}-{mes}', f'Cierre mes {mes} curso {curso}')
            flash('Cierre mensual creado correctamente.', 'success')
        except Exception:
            db.rollback()
            flash('Ya existe un cierre para ese colegio, curso y mes. Elimina el anterior si necesitas recalcularlo.', 'warning')
        return redirect(url_for('cierres_list', mes=mes, colegio_id=colegio_id, curso=curso))

    @app.post('/cierres/<int:cierre_id>/eliminar')
    @role_required('admin')
    def cierres_delete(cierre_id: int):
        db = get_db()
        cierre = db.fetchone('SELECT * FROM cierres_mensuales WHERE id = ?', (cierre_id,))
        if cierre:
            db.execute('DELETE FROM cierres_mensuales WHERE id = ?', (cierre_id,))
            db.commit()
            log_audit('eliminar', 'cierre_mensual', cierre_id, f"Cierre {cierre['mes']} {cierre['curso']}")
            flash('Cierre eliminado.', 'success')
        return redirect(url_for('cierres_list'))

    @app.route('/cierres/<int:cierre_id>/reporte.pdf')
    @login_required
    def cierre_reporte_pdf(cierre_id: int):
        db = get_db()
        cierre = db.fetchone('SELECT cm.*, c.nombre AS colegio_nombre FROM cierres_mensuales cm LEFT JOIN colegios c ON c.id = cm.colegio_id WHERE cm.id = ?', (cierre_id,))
        if not cierre or not ensure_course_access(cierre['curso'], cierre['colegio_id']):
            flash('No tienes acceso a ese cierre.', 'danger')
            return redirect(url_for('cierres_list'))
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=18*mm, bottomMargin=18*mm)
        styles = getSampleStyleSheet()
        data = [
            ['Colegio', cierre['colegio_nombre'] or '-'], ['Curso', cierre['curso']], ['Mes', cierre['mes']],
            ['Ingresos', formato_monto(cierre['ingresos'])], ['Gastos', formato_monto(cierre['gastos'])], ['Saldo', formato_monto(cierre['saldo'])],
            ['Alumnos activos', str(cierre['alumnos_activos'])], ['Pagado cuotas', formato_monto(cierre['total_pagado'])], ['Deuda cuotas', formato_monto(cierre['deuda_total'])],
            ['Creado por', cierre['creado_por'] or '-'], ['Fecha cierre', cierre['creado_en']], ['Observación', cierre['observacion'] or '-'],
        ]
        table = Table(data, colWidths=[45*mm, 115*mm])
        table.setStyle(TableStyle([('BACKGROUND', (0,0), (0,-1), colors.HexColor('#1E3A8A')), ('TEXTCOLOR', (0,0), (0,-1), colors.whitesmoke), ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'), ('GRID', (0,0), (-1,-1), 0.4, colors.grey), ('PADDING', (0,0), (-1,-1), 8)]))
        elements = [Paragraph('<b>ContaCurso</b>', styles['Title']), Paragraph('Reporte de cierre mensual', styles['Heading2']), Spacer(1, 8), table]
        doc.build(elements)
        buffer.seek(0)
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f"cierre_{cierre['mes']}_{cierre['curso']}.pdf")

    @app.route('/auditoria')
    @role_required('admin')
    def auditoria_list():
        db = get_db()
        entidad = request.args.get('entidad', '').strip()
        usuario = request.args.get('usuario', '').strip()
        limit_raw = request.args.get('limit', '200').strip()
        try:
            limit = max(50, min(int(limit_raw), 1000))
        except Exception:
            limit = 200
        sql = 'SELECT * FROM auditoria_acciones WHERE 1=1'
        params: list[Any] = []
        if entidad:
            sql += ' AND entidad = ?'
            params.append(entidad)
        if usuario:
            sql += ' AND lower(username) LIKE ?'
            params.append('%' + usuario.lower() + '%')
        sql += ' ORDER BY fecha DESC, id DESC LIMIT ?'
        params.append(limit)
        acciones = db.fetchall(sql, params)
        entidades = db.fetchall("SELECT DISTINCT entidad FROM auditoria_acciones WHERE entidad IS NOT NULL AND entidad <> '' ORDER BY entidad")
        return render_template('auditoria.html', acciones=acciones, entidades=entidades, entidad=entidad, usuario=usuario, limit=limit)

    @app.route('/usuarios')
    @role_required('admin')
    def usuarios_list():
        db = get_db()
        usuarios = db.fetchall('''
            SELECT u.id, u.username, u.email, u.nombre, u.role, u.curso, u.colegio_id, u.activo,
                   c.nombre AS colegio_nombre
            FROM usuarios u
            LEFT JOIN colegios c ON c.id = u.colegio_id
            ORDER BY u.role, COALESCE(c.nombre, ''), u.nombre, u.username
        ''')
        permisos_por_usuario: dict[int, list[Any]] = {}
        for row in db.fetchall('''
            SELECT urc.*, c.nombre AS colegio_nombre
            FROM usuario_roles_curso urc
            INNER JOIN colegios c ON c.id = urc.colegio_id
            ORDER BY c.nombre, urc.curso, urc.rol_curso
        '''):
            permisos_por_usuario.setdefault(int(row['usuario_id']), []).append(row)
        return render_template('usuarios_list.html', usuarios=usuarios, permisos_por_usuario=permisos_por_usuario)

    def guardar_permisos_usuario(db: DBAdapter, user_id: int, role: str) -> None:
        db.execute('DELETE FROM usuario_roles_curso WHERE usuario_id=?', (user_id,))
        if role in ('admin', 'admin_global'):
            return
        colegios_form = request.form.getlist('perm_colegio_id')
        cursos_form = request.form.getlist('perm_curso')
        roles_form = request.form.getlist('perm_rol')
        for colegio_id, curso, rol_curso in zip(colegios_form, cursos_form, roles_form):
            curso = (curso or '').strip()
            if not colegio_id or not str(colegio_id).isdigit() or not curso or rol_curso not in COURSE_ROLES:
                continue
            db.execute(
                'INSERT INTO usuario_roles_curso (usuario_id, colegio_id, curso, rol_curso) VALUES (?, ?, ?, ?)',
                (user_id, int(colegio_id), curso, rol_curso),
            )

    @app.route('/usuarios/nuevo', methods=['GET', 'POST'])
    @role_required('admin')
    def usuarios_new():
        db = get_db()
        if request.method == 'POST':
            nombre = request.form.get('nombre', '').strip()
            username = request.form.get('username', '').strip().lower()
            email = request.form.get('email', '').strip().lower()
            password = request.form.get('password', '')
            role = request.form.get('role', 'solo_lectura')
            activo = 1 if request.form.get('activo') == 'on' else 0
            if not nombre or not username or not password:
                flash('Nombre, usuario y contraseña son obligatorios.', 'danger')
            elif role not in ALLOWED_ROLES:
                flash('Rol inválido.', 'danger')
            elif db.fetchone('SELECT 1 FROM usuarios WHERE lower(username)=?', (username,)):
                flash('Ese nombre de usuario ya existe.', 'danger')
            elif email and db.fetchone("SELECT 1 FROM usuarios WHERE lower(COALESCE(email,''))=?", (email,)):
                flash('Ese correo ya existe.', 'danger')
            else:
                db.execute(
                    'INSERT INTO usuarios (username, email, password_hash, role, nombre, curso, colegio_id, activo) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                    (username, email or None, generate_password_hash(password), role, nombre, None, None if role in ('admin', 'admin_global') else 1, activo)
                )
                new_user = db.fetchone('SELECT id FROM usuarios WHERE lower(username)=?', (username,))
                if new_user:
                    guardar_permisos_usuario(db, int(new_user['id']), role)
                db.commit()
                flash('Usuario creado.', 'success')
                return redirect(url_for('usuarios_list'))
        return render_template('usuarios_form.html', usuario=None, roles=ALLOWED_ROLES, permisos=[])

    @app.route('/usuarios/<int:user_id>/editar', methods=['GET', 'POST'])
    @role_required('admin')
    def usuarios_edit(user_id: int):
        db = get_db()
        usuario = db.fetchone('SELECT id, username, email, role, nombre, curso, colegio_id, activo FROM usuarios WHERE id=?', (user_id,))
        if not usuario:
            flash('Usuario no encontrado.', 'danger')
            return redirect(url_for('usuarios_list'))
        if request.method == 'POST':
            nombre = request.form.get('nombre', '').strip()
            username = request.form.get('username', '').strip().lower()
            password = request.form.get('password', '')
            role = request.form.get('role', 'solo_lectura')
            activo = 1 if request.form.get('activo') == 'on' else 0
            if not nombre or not username:
                flash('Nombre y usuario son obligatorios.', 'danger')
            elif role not in ALLOWED_ROLES:
                flash('Rol inválido.', 'danger')
            elif db.fetchone('SELECT 1 FROM usuarios WHERE lower(username)=? AND id<>?', (username, user_id)):
                flash('Ese nombre de usuario ya existe.', 'danger')
            elif email and db.fetchone("SELECT 1 FROM usuarios WHERE lower(COALESCE(email,''))=? AND id<>?", (email, user_id)):
                flash('Ese correo ya existe.', 'danger')
            else:
                colegio_id = None if role in ('admin', 'admin_global') else 1
                if password:
                    db.execute('UPDATE usuarios SET nombre=?, username=?, email=?, role=?, curso=?, colegio_id=?, activo=?, password_hash=? WHERE id=?',
                               (nombre, username, email or None, role, None, colegio_id, activo, generate_password_hash(password), user_id))
                else:
                    db.execute('UPDATE usuarios SET nombre=?, username=?, email=?, role=?, curso=?, colegio_id=?, activo=? WHERE id=?',
                               (nombre, username, email or None, role, None, colegio_id, activo, user_id))
                guardar_permisos_usuario(db, user_id, role)
                db.commit()
                flash('Usuario actualizado.', 'success')
                return redirect(url_for('usuarios_list'))
        permisos = db.fetchall('SELECT * FROM usuario_roles_curso WHERE usuario_id=? ORDER BY colegio_id, curso, rol_curso', (user_id,))
        return render_template('usuarios_form.html', usuario=usuario, roles=ALLOWED_ROLES, permisos=permisos)

    @app.post('/usuarios/<int:user_id>/eliminar')
    @role_required('admin')
    def usuarios_delete(user_id: int):
        if int(current_user.id) == user_id:
            flash('No puedes eliminar tu propio usuario.', 'danger')
            return redirect(url_for('usuarios_list'))
        db = get_db()
        db.execute('DELETE FROM usuario_roles_curso WHERE usuario_id=?', (user_id,))
        db.execute('DELETE FROM usuarios WHERE id=?', (user_id,))
        db.commit()
        flash('Usuario eliminado.', 'success')
        return redirect(url_for('usuarios_list'))

    @app.route('/alumnos')
    @login_required
    def alumnos_list():
        db = get_db()
        q = request.args.get('q', '').strip()
        mes = request.args.get('mes') or datetime.today().strftime('%Y-%m')
        sql = """
            SELECT a.id, a.nombre, a.curso, a.colegio_id, c.nombre AS colegio_nombre, a.cuota_mensual, a.activo,
                   COALESCE(SUM(CASE WHEN p.mes = ? THEN p.monto ELSE 0 END), 0) AS pagado_mes
            FROM alumnos a
            LEFT JOIN pagos_alumnos p ON p.alumno_id = a.id
            LEFT JOIN colegios c ON c.id = a.colegio_id
            WHERE 1=1
        """
        params: list[Any] = [mes]
        sql, params = course_filter_sql(sql, params, 'a')
        if q:
            sql += " AND (LOWER(COALESCE(a.nombre, '')) LIKE ? OR LOWER(COALESCE(a.curso, '')) LIKE ?)"
            like = sql_like_ci(q)
            params.extend([like, like])
        sql += ' GROUP BY a.id, a.nombre, a.curso, a.colegio_id, c.nombre, a.cuota_mensual, a.activo ORDER BY a.curso, a.nombre'
        alumnos = db.fetchall(sql, params)
        deuda_total = sum(max(float(a['cuota_mensual']) - float(a['pagado_mes']), 0) for a in alumnos if a['activo'])
        return render_template('alumnos_list.html', alumnos=alumnos, q=q, mes=mes, deuda_total=deuda_total)

    @app.route('/alumnos/nuevo', methods=['GET', 'POST'])
    @role_required('admin', 'presidente', 'tesorero', 'secretario')
    def alumnos_new():
        db = get_db()
        if request.method == 'POST':
            nombre = request.form.get('nombre', '').strip()
            curso = (request.form.get('curso', '').strip() or user_course_scope() or '').strip()
            cuota = parse_float(request.form.get('cuota_mensual', '0'))
            apoderado = request.form.get('apoderado', '').strip()
            telefono = request.form.get('telefono', '').strip()
            email = request.form.get('email', '').strip().lower()
            direccion = request.form.get('direccion', '').strip()
            observacion_ficha = request.form.get('observacion_ficha', '').strip()
            activo = 1 if request.form.get('activo') == 'on' else 0
            if not nombre:
                flash('El nombre es obligatorio.', 'danger')
            elif not ensure_course_access(curso, resolve_colegio_for_course(curso)):
                flash('No puedes crear alumnos en otro curso.', 'danger')
            elif alumno_duplicado(db, nombre, curso, colegio_id=resolve_colegio_for_course(curso)):
                flash('Ya existe un alumno con ese nombre y curso.', 'danger')
            else:
                db.execute(
                    'INSERT INTO alumnos (nombre, curso, colegio_id, cuota_mensual, apoderado, telefono, email, direccion, observacion_ficha, activo) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                    (nombre, curso, resolve_colegio_for_course(curso), cuota, apoderado, telefono, email or None, direccion, observacion_ficha, activo),
                )
                db.commit()
                log_audit('crear', 'alumno', db.fetchone('SELECT MAX(id) AS id FROM alumnos')['id'], nombre)
                flash('Alumno creado correctamente.', 'success')
                return redirect(url_for('alumnos_list'))
        return render_template('alumnos_form.html', alumno=None)

    @app.route('/alumnos/<int:alumno_id>/editar', methods=['GET', 'POST'])
    @role_required('admin', 'presidente', 'tesorero', 'secretario')
    def alumnos_edit(alumno_id: int):
        db = get_db()
        alumno = fetch_alumno_permitido(db, alumno_id)
        if not alumno:
            flash('Alumno no encontrado.', 'danger')
            return redirect(url_for('alumnos_list'))
        if request.method == 'POST':
            nombre = request.form.get('nombre', '').strip()
            curso = (request.form.get('curso', '').strip() or user_course_scope() or '').strip()
            cuota = parse_float(request.form.get('cuota_mensual', '0'))
            apoderado = request.form.get('apoderado', '').strip()
            telefono = request.form.get('telefono', '').strip()
            email = request.form.get('email', '').strip().lower()
            direccion = request.form.get('direccion', '').strip()
            observacion_ficha = request.form.get('observacion_ficha', '').strip()
            activo = 1 if request.form.get('activo') == 'on' else 0
            if not nombre:
                flash('El nombre es obligatorio.', 'danger')
            elif not ensure_course_access(curso, resolve_colegio_for_course(curso)):
                flash('No puedes mover alumnos a otro curso.', 'danger')
            elif alumno_duplicado(db, nombre, curso, exclude_id=alumno_id, colegio_id=resolve_colegio_for_course(curso)):
                flash('Ya existe otro alumno con ese nombre y curso.', 'danger')
            else:
                db.execute(
                    'UPDATE alumnos SET nombre = ?, curso = ?, colegio_id = ?, cuota_mensual = ?, apoderado = ?, telefono = ?, email = ?, direccion = ?, observacion_ficha = ?, activo = ? WHERE id = ?',
                    (nombre, curso, resolve_colegio_for_course(curso), cuota, apoderado, telefono, email or None, direccion, observacion_ficha, activo, alumno_id),
                )
                db.commit()
                flash('Alumno actualizado.', 'success')
                return redirect(url_for('alumnos_list'))
        return render_template('alumnos_form.html', alumno=alumno)

    @app.post('/alumnos/<int:alumno_id>/eliminar')
    @role_required('admin', 'presidente', 'tesorero', 'secretario')
    def alumnos_delete(alumno_id: int):
        db = get_db()
        alumno = fetch_alumno_permitido(db, alumno_id)
        if not alumno:
            flash('Alumno no encontrado.', 'danger')
            return redirect(url_for('alumnos_list'))
        db.execute('DELETE FROM pagos_alumnos WHERE alumno_id = ?', (alumno_id,))
        db.execute('DELETE FROM alumnos WHERE id = ?', (alumno_id,))
        db.commit()
        flash(f'Alumno eliminado: {alumno["nombre"]}.', 'success')
        return redirect(url_for('alumnos_list'))


    @app.route('/alumnos/exportar/excel')
    @login_required
    def alumnos_export_excel():
        db = get_db()
        mes = request.args.get('mes') or datetime.today().strftime('%Y-%m')
        sql = """
            SELECT a.nombre, a.curso, c.nombre AS colegio, a.apoderado, a.telefono, a.cuota_mensual, a.activo,
                   COALESCE(SUM(CASE WHEN p.mes = ? THEN p.monto ELSE 0 END), 0) AS pagado_mes
            FROM alumnos a
            LEFT JOIN colegios c ON c.id = COALESCE(a.colegio_id, 1)
            LEFT JOIN pagos_alumnos p ON p.alumno_id = a.id
            WHERE 1=1
        """
        params = [mes]
        sql, params = course_filter_sql(sql, params, 'a')
        sql += ' GROUP BY a.id, a.nombre, a.curso, c.nombre, a.apoderado, a.telefono, a.cuota_mensual, a.activo ORDER BY c.nombre, a.curso, a.nombre'
        rows = db.fetchall(sql, params)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Alumnos'
        ws.append(['Colegio', 'Curso', 'Alumno', 'Apoderado', 'Teléfono', 'Cuota mensual', 'Pagado mes', 'Deuda mes', 'Estado'])
        for r in rows:
            cuota = float(r['cuota_mensual'] or 0)
            pagado = float(r['pagado_mes'] or 0)
            deuda = max(cuota - pagado, 0) if r['activo'] else 0
            estado, _ = estado_cuota(cuota, pagado) if r['activo'] else ('Inactivo', '')
            ws.append([r['colegio'], r['curso'], r['nombre'], r['apoderado'], r['telefono'], cuota, pagado, deuda, estado])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(bio, as_attachment=True, download_name=f'alumnos_{mes}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/alumnos/<int:alumno_id>/ficha.pdf')
    @login_required
    def alumno_ficha_pdf(alumno_id: int):
        db = get_db()
        alumno = fetch_alumno_permitido(db, alumno_id)
        if not alumno:
            flash('Alumno no encontrado.', 'danger')
            return redirect(url_for('alumnos_list'))
        pagos = db.fetchall('SELECT fecha, mes, monto, observacion FROM pagos_alumnos WHERE alumno_id=? ORDER BY fecha DESC, id DESC LIMIT 20', (alumno_id,))
        movimientos = db.fetchall('SELECT fecha, tipo, concepto, monto FROM movimientos WHERE alumno_id=? ORDER BY fecha DESC, id DESC LIMIT 20', (alumno_id,))
        bio = BytesIO()
        doc = SimpleDocTemplate(bio, pagesize=A4, rightMargin=14*mm, leftMargin=14*mm, topMargin=14*mm, bottomMargin=14*mm)
        styles = getSampleStyleSheet()
        story = [Paragraph(f'{APP_NAME} - Ficha del alumno', styles['Title']), Spacer(1, 6*mm)]
        info = [
            ['Colegio', alumno['colegio_nombre'] or '-'], ['Curso', alumno['curso'] or '-'], ['Alumno', alumno['nombre']],
            ['Apoderado', alumno['apoderado'] or '-'], ['Teléfono', alumno['telefono'] or '-'], ['Cuota mensual', formato_monto(alumno['cuota_mensual'])],
        ]
        t = Table(info, colWidths=[40*mm, 120*mm])
        t.setStyle(TableStyle([('GRID',(0,0),(-1,-1),0.25,colors.grey),('BACKGROUND',(0,0),(0,-1),colors.lightgrey)]))
        story += [t, Spacer(1, 8*mm), Paragraph('Últimos pagos', styles['Heading2'])]
        data = [['Fecha','Mes','Monto','Observación']] + [[p['fecha'], p['mes'], formato_monto(p['monto']), p['observacion'] or ''] for p in pagos]
        if len(data) == 1: data.append(['-','-','-','Sin pagos'])
        tp = Table(data, repeatRows=1, colWidths=[28*mm,25*mm,30*mm,85*mm])
        tp.setStyle(TableStyle([('GRID',(0,0),(-1,-1),0.25,colors.grey),('BACKGROUND',(0,0),(-1,0),colors.lightgrey)]))
        story += [tp, Spacer(1, 8*mm), Paragraph('Últimos movimientos asociados', styles['Heading2'])]
        data2 = [['Fecha','Tipo','Concepto','Monto']] + [[m['fecha'], m['tipo'], m['concepto'], formato_monto(m['monto'])] for m in movimientos]
        if len(data2) == 1: data2.append(['-','-','Sin movimientos','-'])
        tm = Table(data2, repeatRows=1, colWidths=[28*mm,25*mm,95*mm,30*mm])
        tm.setStyle(TableStyle([('GRID',(0,0),(-1,-1),0.25,colors.grey),('BACKGROUND',(0,0),(-1,0),colors.lightgrey)]))
        story.append(tm)
        doc.build(story)
        bio.seek(0)
        return send_file(bio, as_attachment=True, download_name=f'ficha_{secure_filename(alumno["nombre"])}.pdf', mimetype='application/pdf')

    @app.route('/alumnos/<int:alumno_id>')
    @login_required
    def alumno_detail(alumno_id: int):
        db = get_db()
        alumno = fetch_alumno_permitido(db, alumno_id)
        if not alumno:
            flash('Alumno no encontrado.', 'danger')
            return redirect(url_for('alumnos_list'))
        mes_actual = request.args.get('mes') or datetime.today().strftime('%Y-%m')
        resumen_mes = db.fetchone(
            """
            SELECT COALESCE(SUM(CASE WHEN p.mes = ? THEN p.monto ELSE 0 END), 0) AS pagado_mes,
                   COUNT(CASE WHEN p.mes = ? THEN 1 END) AS pagos_mes
            FROM pagos_alumnos p
            WHERE p.alumno_id = ?
            """,
            (mes_actual, mes_actual, alumno_id),
        )
        resumen_aportes = db.fetchone(
            """
            SELECT COALESCE(SUM(CASE WHEN m.tipo = 'ingreso' THEN m.monto ELSE 0 END), 0) AS ingresos_actividad,
                   COALESCE(SUM(CASE WHEN m.tipo = 'gasto' THEN m.monto ELSE 0 END), 0) AS gastos_asociados,
                   COUNT(*) AS movimientos_asociados
            FROM movimientos m
            WHERE m.alumno_id = ?
            """,
            (alumno_id,),
        )
        historial_cuotas = db.fetchall(
            """
            SELECT p.id, p.fecha, p.mes, p.monto, p.observacion, 'Cuota mensual' AS tipo, NULL AS actividad
            FROM pagos_alumnos p
            WHERE p.alumno_id = ?
            ORDER BY p.fecha DESC, p.id DESC
            """,
            (alumno_id,),
        )
        historial_aportes = db.fetchall(
            """
            SELECT m.id, m.fecha, substr(m.fecha,1,7) AS mes, m.monto, m.observacion,
                   CASE WHEN m.tipo = 'gasto' THEN 'Egreso asociado' ELSE 'Aporte actividad' END AS tipo,
                   COALESCE(a.nombre, '-') AS actividad
            FROM movimientos m
            LEFT JOIN actividades a ON a.id = m.actividad_id
            WHERE m.alumno_id = ?
               OR (m.origen = 'actividad_alumno' AND m.alumno_id IS NULL AND LOWER(m.concepto) LIKE ?)
            ORDER BY m.fecha DESC, m.id DESC
            """,
            (alumno_id, sql_like_ci(f'Aporte actividad alumno: {alumno["nombre"]}')[:-1] + '%'),
        )
        historial = sorted([dict(x) for x in historial_cuotas] + [dict(x) for x in historial_aportes], key=lambda x: (x['fecha'], x['id']), reverse=True)
        actividad_resumen = db.fetchall(
            """
            SELECT COALESCE(a.id, 0) AS actividad_id, COALESCE(a.nombre, 'Sin actividad') AS actividad,
                   COALESCE(SUM(CASE WHEN m.tipo = 'ingreso' THEN m.monto ELSE 0 END), 0) AS ingresos,
                   COALESCE(SUM(CASE WHEN m.tipo = 'gasto' THEN m.monto ELSE 0 END), 0) AS egresos,
                   COUNT(*) AS movimientos
            FROM movimientos m
            LEFT JOIN actividades a ON a.id = m.actividad_id
            WHERE m.alumno_id = ?
            GROUP BY a.id, a.nombre
            ORDER BY actividad
            """,
            (alumno_id,),
        )
        deuda_mes = max(float(alumno['cuota_mensual']) - float(resumen_mes['pagado_mes'] or 0), 0) if alumno['activo'] else 0
        resumen = {
            'mes': mes_actual,
            'pagado_mes': float(resumen_mes['pagado_mes'] or 0),
            'pagos_mes': int(resumen_mes['pagos_mes'] or 0),
            'deuda_mes': deuda_mes,
            'ingresos_actividad': float(resumen_aportes['ingresos_actividad'] or 0),
            'gastos_asociados': float(resumen_aportes['gastos_asociados'] or 0),
            'movimientos_asociados': int(resumen_aportes['movimientos_asociados'] or 0),
        }
        ficha = {
            'apoderado': alumno.get('apoderado') if hasattr(alumno, 'get') else alumno['apoderado'],
            'telefono': alumno.get('telefono') if hasattr(alumno, 'get') else alumno['telefono'],
            'direccion': alumno.get('direccion') if hasattr(alumno, 'get') else alumno['direccion'],
            'observacion_ficha': alumno.get('observacion_ficha') if hasattr(alumno, 'get') else alumno['observacion_ficha'],
        }
        return render_template('alumno_detail.html', alumno=alumno, historial=historial, resumen=resumen, actividad_resumen=actividad_resumen, ficha=ficha)

    @app.route('/pagos')
    @login_required
    def pagos_list():
        db = get_db()
        mes = request.args.get('mes', '').strip()
        sql = """
            SELECT p.id, p.folio, p.alumno_id, a.nombre, a.curso, COALESCE(a.email, '') AS email, p.fecha, p.mes, p.monto, p.observacion, p.movimiento_id
            FROM pagos_alumnos p
            INNER JOIN alumnos a ON a.id = p.alumno_id
            WHERE 1=1
        """
        params: list[Any] = []
        sql, params = course_filter_sql(sql, params, 'a')
        if mes:
            sql += ' AND p.mes = ?'
            params.append(mes)
        sql += ' ORDER BY p.fecha DESC, a.nombre'
        pagos = db.fetchall(sql, params)
        return render_template('pagos_list.html', pagos=pagos, mes=mes)

    @app.route('/pagos/nuevo', methods=['GET', 'POST'])
    @role_required('admin', 'presidente', 'tesorero', 'secretario')
    def pagos_new():
        db = get_db()

        selected_pago_colegio_id = None
        raw_colegio = (request.form.get('colegio_id') if request.method == 'POST' else request.args.get('colegio_id')) or ''
        if str(raw_colegio).strip().isdigit():
            selected_pago_colegio_id = int(str(raw_colegio).strip())

        alumnos_sql = """
            SELECT a.id, a.nombre, a.curso, a.colegio_id, a.cuota_mensual, c.nombre AS colegio_nombre
            FROM alumnos a
            LEFT JOIN colegios c ON c.id = a.colegio_id
            WHERE a.activo = 1
        """
        alumnos_params: list[Any] = []
        if current_user.is_admin_global():
            if selected_pago_colegio_id:
                alumnos_sql += ' AND COALESCE(a.colegio_id, 1) = ?'
                alumnos_params.append(selected_pago_colegio_id)
            curso_admin = admin_selected_course()
            if curso_admin:
                alumnos_sql += " AND lower(trim(COALESCE(a.curso, ''))) = lower(trim(?))"
                alumnos_params.append(curso_admin)
        else:
            alumnos_sql, alumnos_params = course_filter_sql(alumnos_sql, alumnos_params, 'a')
        alumnos_sql += ' ORDER BY c.nombre, a.curso, a.nombre'
        alumnos = db.fetchall(alumnos_sql, alumnos_params)

        actividades_sql = """
            SELECT a.id, a.nombre, a.fecha, a.curso, a.colegio_id, c.nombre AS colegio_nombre
            FROM actividades a
            LEFT JOIN colegios c ON c.id = a.colegio_id
            WHERE 1=1
        """
        actividades_params: list[Any] = []
        if current_user.is_admin_global():
            if selected_pago_colegio_id:
                actividades_sql += ' AND COALESCE(a.colegio_id, 1) = ?'
                actividades_params.append(selected_pago_colegio_id)
            curso_admin = admin_selected_course()
            if curso_admin:
                actividades_sql += " AND lower(trim(COALESCE(a.curso, ''))) = lower(trim(?))"
                actividades_params.append(curso_admin)
        else:
            actividades_sql, actividades_params = course_filter_sql(actividades_sql, actividades_params, 'a')
        actividades_sql += ' ORDER BY c.nombre, a.fecha DESC, a.nombre'
        actividades = db.fetchall(actividades_sql, actividades_params)

        if request.method == 'POST':
            alumno_id = int(request.form.get('alumno_id', '0') or 0)
            fecha = request.form.get('fecha', '').strip()
            mes = request.form.get('mes', '').strip()
            monto = parse_float(request.form.get('monto', '0'))
            observacion = request.form.get('observacion', '').strip()
            tipo_pago = request.form.get('tipo_pago', 'cuota_mensual')
            actividad_raw = request.form.get('actividad_id', '').strip()
            actividad_id = int(actividad_raw) if actividad_raw else None
            try:
                validar_fecha(fecha)
                datetime.strptime(mes + '-01', '%Y-%m-%d')
            except Exception:
                flash('Fecha o mes inválido.', 'danger')
                return render_template('pagos_form.html', alumnos=alumnos, actividades=actividades, pago=None, selected_pago_colegio_id=selected_pago_colegio_id)
            alumno_perm = fetch_alumno_permitido(db, alumno_id) if alumno_id else None
            if not alumno_id:
                flash('Debes seleccionar un alumno.', 'danger')
            elif monto <= 0:
                flash('El monto debe ser mayor a cero.', 'danger')
            elif tipo_pago == 'actividad_alumno' and not actividad_id:
                flash('Debes seleccionar una actividad para un aporte.', 'danger')
            elif tipo_pago == 'cuota_mensual' and pago_duplicado(db, alumno_id, mes):
                flash('Ese alumno ya tiene un pago registrado para ese mes.', 'danger')
            elif not alumno_perm:
                flash('Alumno no encontrado para el colegio/curso seleccionado.', 'danger')
            elif mes_esta_cerrado(db, alumno_perm['colegio_id'], alumno_perm['curso'], mes) and not current_user.is_admin_global():
                flash('El mes está cerrado para ese colegio y curso. Solo admin puede modificarlo.', 'danger')
            else:
                registrar_pago_alumno(db, alumno_id, fecha, mes, monto, observacion, actividad_id, tipo_pago)
                db.commit()
                log_audit('crear', 'pago', alumno_id, f'Mes {mes}, monto {monto}')
                flash('Pago registrado correctamente.', 'success')
                return redirect(url_for('pagos_list', colegio_id=selected_pago_colegio_id) if selected_pago_colegio_id else url_for('pagos_list'))
        return render_template('pagos_form.html', alumnos=alumnos, actividades=actividades, pago=None, selected_pago_colegio_id=selected_pago_colegio_id)

    @app.route('/pagos/<int:pago_id>/editar', methods=['GET', 'POST'])
    @role_required('admin', 'presidente', 'tesorero', 'secretario')
    def pagos_edit(pago_id: int):
        db = get_db()
        pago = fetch_pago_permitido(db, pago_id)
        if not pago:
            flash('Pago no encontrado.', 'danger')
            return redirect(url_for('pagos_list'))
        alumnos_sql = 'SELECT a.id, a.nombre, a.curso, a.colegio_id, a.cuota_mensual, c.nombre AS colegio_nombre FROM alumnos a LEFT JOIN colegios c ON c.id = a.colegio_id WHERE (a.activo = 1 OR a.id = ?)'
        alumnos_params: list[Any] = [pago['alumno_id']]
        alumnos_sql, alumnos_params = course_filter_sql(alumnos_sql, alumnos_params, 'a')
        alumnos_sql += ' ORDER BY a.curso, a.nombre'
        alumnos = db.fetchall(alumnos_sql, alumnos_params)
        actividades_sql = 'SELECT a.id, a.nombre, a.fecha, a.curso, a.colegio_id, c.nombre AS colegio_nombre FROM actividades a LEFT JOIN colegios c ON c.id = a.colegio_id WHERE 1=1'
        actividades_params: list[Any] = []
        actividades_sql, actividades_params = course_filter_sql(actividades_sql, actividades_params, 'a')
        actividades_sql += ' ORDER BY fecha DESC, nombre'
        actividades = db.fetchall(actividades_sql, actividades_params)
        if request.method == 'POST':
            alumno_id = int(request.form.get('alumno_id', '0') or 0)
            fecha = request.form.get('fecha', '').strip()
            mes = request.form.get('mes', '').strip()
            monto = parse_float(request.form.get('monto', '0'))
            observacion = request.form.get('observacion', '').strip()
            try:
                validar_fecha(fecha)
                datetime.strptime(mes + '-01', '%Y-%m-%d')
            except Exception:
                flash('Fecha o mes inválido.', 'danger')
                return render_template('pagos_form.html', alumnos=alumnos, actividades=actividades, pago=pago)
            alumno = fetch_alumno_permitido(db, alumno_id)
            if not alumno:
                flash('Alumno no encontrado para tu curso.', 'danger')
            elif monto <= 0:
                flash('El monto debe ser mayor a cero.', 'danger')
            elif mes_esta_cerrado(db, pago['colegio_id'], pago['curso_alumno'], pago['mes']) and not current_user.is_admin_global():
                flash('El mes original está cerrado. Solo admin puede modificarlo.', 'danger')
            elif mes_esta_cerrado(db, alumno['colegio_id'], alumno['curso'], mes) and not current_user.is_admin_global():
                flash('El mes nuevo está cerrado. Solo admin puede modificarlo.', 'danger')
            elif db.fetchone('SELECT 1 FROM pagos_alumnos WHERE alumno_id=? AND mes=? AND id<>?', (alumno_id, mes, pago_id)):
                flash('Ese alumno ya tiene otro pago registrado para ese mes.', 'danger')
            else:
                db.execute('UPDATE pagos_alumnos SET alumno_id=?, fecha=?, mes=?, monto=?, observacion=? WHERE id=?',
                           (alumno_id, fecha, mes, monto, observacion, pago_id))
                db.execute('UPDATE movimientos SET fecha=?, concepto=?, monto=?, alumno_id=?, observacion=?, curso=?, colegio_id=? WHERE id=?',
                           (fecha, f'Cuota mensual alumno: {obtener_nombre_alumno(db, alumno_id)} ({mes})', monto, alumno_id, observacion, alumno['curso'], alumno['colegio_id'] if 'colegio_id' in alumno.keys() else 1, pago['movimiento_id']))
                db.commit()
                flash('Pago actualizado.', 'success')
                return redirect(url_for('pagos_list'))
        return render_template('pagos_form.html', alumnos=alumnos, actividades=actividades, pago=pago)

    @app.post('/pagos/<int:pago_id>/eliminar')
    @role_required('admin', 'presidente', 'tesorero', 'secretario')
    def pagos_delete(pago_id: int):
        db = get_db()
        pago = fetch_pago_permitido(db, pago_id)
        if not pago:
            flash('Pago no encontrado.', 'danger')
            return redirect(url_for('pagos_list'))
        if mes_esta_cerrado(db, pago['colegio_id'], pago['curso_alumno'], pago['mes']) and not current_user.is_admin_global():
            flash('El mes está cerrado para este pago. Solo admin puede eliminarlo.', 'danger')
            return redirect(url_for('pagos_list'))
        db.execute('DELETE FROM pagos_alumnos WHERE id=?', (pago_id,))
        if pago['movimiento_id']:
            db.execute('DELETE FROM movimientos WHERE id=?', (pago['movimiento_id'],))
        db.commit()
        flash('Pago eliminado.', 'success')
        return redirect(url_for('pagos_list'))

    @app.route('/movimientos')
    @login_required
    def movimientos_list():
        db = get_db()
        tipo = request.args.get('tipo', 'Todos')
        mes = request.args.get('mes', '')
        q = request.args.get('q', '').strip()
        fecha_desde = request.args.get('fecha_desde', '').strip()
        fecha_hasta = request.args.get('fecha_hasta', '').strip()
        actividad_id = request.args.get('actividad_id', '').strip()
        alumno_id = request.args.get('alumno_id', '').strip()
        movimientos = obtener_movimientos_filtrados(db, tipo=tipo, mes=mes, q=q, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, actividad_id=actividad_id, alumno_id=alumno_id, curso_scope=current_course_filter(), colegio_scope=selected_colegio_id() if current_user.is_admin_global() else getattr(current_user, 'colegio_id', None))
        actividades_sql = 'SELECT id, nombre, fecha, curso FROM actividades a WHERE 1=1'
        actividades_params: list[Any] = []
        actividades_sql, actividades_params = course_filter_sql(actividades_sql, actividades_params, 'a')
        actividades_sql += ' ORDER BY fecha DESC, nombre'
        actividades = db.fetchall(actividades_sql, actividades_params)
        alumnos_sql = 'SELECT id, nombre, curso FROM alumnos a WHERE activo = 1'
        alumnos_params: list[Any] = []
        alumnos_sql, alumnos_params = course_filter_sql(alumnos_sql, alumnos_params, 'a')
        alumnos_sql += ' ORDER BY a.curso, a.nombre'
        alumnos = db.fetchall(alumnos_sql, alumnos_params)
        return render_template('movimientos_list.html', movimientos=movimientos, tipo=tipo, mes=mes, q=q, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, actividad_id=actividad_id, alumno_id=alumno_id, actividades=actividades, alumnos=alumnos)

    @app.route('/movimientos/nuevo', methods=['GET', 'POST'])
    @role_required('admin', 'presidente', 'tesorero', 'secretario')
    def movimientos_new():
        db = get_db()
        actividades_sql = 'SELECT id, nombre, fecha, curso FROM actividades a WHERE 1=1'
        actividades_params: list[Any] = []
        actividades_sql, actividades_params = course_filter_sql(actividades_sql, actividades_params, 'a')
        actividades_sql += ' ORDER BY fecha DESC, nombre'
        actividades = db.fetchall(actividades_sql, actividades_params)
        alumnos_sql = 'SELECT id, nombre, curso FROM alumnos a WHERE activo = 1'
        alumnos_params: list[Any] = []
        alumnos_sql, alumnos_params = course_filter_sql(alumnos_sql, alumnos_params, 'a')
        alumnos_sql += ' ORDER BY a.curso, a.nombre'
        alumnos = db.fetchall(alumnos_sql, alumnos_params)
        if request.method == 'POST':
            fecha = request.form.get('fecha', '').strip()
            tipo = request.form.get('tipo', 'ingreso').strip()
            concepto = request.form.get('concepto', '').strip()
            monto = parse_float(request.form.get('monto', '0'))
            actividad_raw = request.form.get('actividad_id', '').strip()
            actividad_id = int(actividad_raw) if actividad_raw else None
            alumno_raw = request.form.get('alumno_id', '').strip()
            alumno_id = int(alumno_raw) if alumno_raw else None
            observacion = request.form.get('observacion', '').strip()
            curso_form = request.form.get('curso', '').strip()
            try:
                validar_fecha(fecha)
            except Exception:
                flash('Fecha inválida.', 'danger')
                return render_template('movimientos_form.html', actividades=actividades, alumnos=alumnos, movimiento=None)
            try:
                curso = resolver_curso_operacion(db, alumno_id, actividad_id, curso_form)
            except ValueError as exc:
                flash(str(exc), 'danger')
                return render_template('movimientos_form.html', actividades=actividades, alumnos=alumnos, movimiento=None)
            colegio_op = resolve_colegio_for_course(curso)
            if monto <= 0:
                flash('El monto debe ser mayor a cero.', 'danger')
                return render_template('movimientos_form.html', actividades=actividades, alumnos=alumnos, movimiento=None)
            try:
                assert_mes_abierto(db, colegio_op, curso, fecha)
            except ValueError as exc:
                flash(str(exc), 'danger')
                return render_template('movimientos_form.html', actividades=actividades, alumnos=alumnos, movimiento=None)
            db.execute(
                'INSERT INTO movimientos (fecha, tipo, concepto, monto, actividad_id, alumno_id, observacion, origen, curso, colegio_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                (fecha, tipo, concepto, monto, actividad_id, alumno_id, observacion, 'general', curso, colegio_op),
            )
            db.commit()
            flash('Movimiento creado.', 'success')
            next_url = request.form.get('next', '').strip()
            return redirect(next_url or url_for('movimientos_list'))
        return render_template('movimientos_form.html', actividades=actividades, alumnos=alumnos, movimiento=None, next_url=request.args.get('next', ''), selected_alumno_id=request.args.get('alumno_id', ''))

    @app.route('/movimientos/<int:movimiento_id>/editar', methods=['GET', 'POST'])
    @role_required('admin', 'presidente', 'tesorero', 'secretario')
    def movimientos_edit(movimiento_id: int):
        db = get_db()
        movimiento = fetch_movimiento_permitido(db, movimiento_id)
        if not movimiento:
            flash('Movimiento no encontrado.', 'danger')
            return redirect(url_for('movimientos_list'))
        actividades_sql = 'SELECT id, nombre, fecha, curso FROM actividades a WHERE 1=1'
        actividades_params: list[Any] = []
        actividades_sql, actividades_params = course_filter_sql(actividades_sql, actividades_params, 'a')
        actividades_sql += ' ORDER BY fecha DESC, nombre'
        actividades = db.fetchall(actividades_sql, actividades_params)
        alumnos_sql = 'SELECT id, nombre, curso FROM alumnos a WHERE (activo = 1 OR id = ?)'
        alumnos_params: list[Any] = [movimiento['alumno_id'] or 0]
        alumnos_sql, alumnos_params = course_filter_sql(alumnos_sql, alumnos_params, 'a')
        alumnos_sql += ' ORDER BY a.curso, a.nombre'
        alumnos = db.fetchall(alumnos_sql, alumnos_params)
        if request.method == 'POST':
            fecha = request.form.get('fecha', '').strip()
            tipo = request.form.get('tipo', 'ingreso').strip()
            concepto = request.form.get('concepto', '').strip()
            monto = parse_float(request.form.get('monto', '0'))
            actividad_raw = request.form.get('actividad_id', '').strip()
            actividad_id = int(actividad_raw) if actividad_raw else None
            alumno_raw = request.form.get('alumno_id', '').strip()
            alumno_id = int(alumno_raw) if alumno_raw else None
            observacion = request.form.get('observacion', '').strip()
            curso_form = request.form.get('curso', '').strip()
            try:
                validar_fecha(fecha)
            except Exception:
                flash('Fecha inválida.', 'danger')
                return render_template('movimientos_form.html', actividades=actividades, alumnos=alumnos, movimiento=movimiento)
            try:
                curso = resolver_curso_operacion(db, alumno_id, actividad_id, curso_form)
            except ValueError as exc:
                flash(str(exc), 'danger')
                return render_template('movimientos_form.html', actividades=actividades, alumnos=alumnos, movimiento=movimiento)
            colegio_op = resolve_colegio_for_course(curso)
            if monto <= 0:
                flash('El monto debe ser mayor a cero.', 'danger')
                return render_template('movimientos_form.html', actividades=actividades, alumnos=alumnos, movimiento=movimiento)
            try:
                assert_mes_abierto(db, movimiento['colegio_ref'], movimiento['curso_ref'], movimiento['fecha'])
                assert_mes_abierto(db, colegio_op, curso, fecha)
            except ValueError as exc:
                flash(str(exc), 'danger')
                return render_template('movimientos_form.html', actividades=actividades, alumnos=alumnos, movimiento=movimiento)
            db.execute('UPDATE movimientos SET fecha=?, tipo=?, concepto=?, monto=?, actividad_id=?, alumno_id=?, observacion=?, curso=?, colegio_id=? WHERE id=?',
                       (fecha, tipo, concepto, monto, actividad_id, alumno_id, observacion, curso, colegio_op, movimiento_id))
            db.commit()
            flash('Movimiento actualizado.', 'success')
            next_url = request.form.get('next', '').strip()
            return redirect(next_url or url_for('movimientos_list'))
        return render_template('movimientos_form.html', actividades=actividades, alumnos=alumnos, movimiento=movimiento, next_url=request.args.get('next', ''), selected_alumno_id=request.args.get('alumno_id', ''))

    @app.post('/movimientos/<int:movimiento_id>/eliminar')
    @role_required('admin', 'presidente', 'tesorero', 'secretario')
    def movimientos_delete(movimiento_id: int):
        db = get_db()
        movimiento = fetch_movimiento_permitido(db, movimiento_id)
        if not movimiento:
            flash('Movimiento no encontrado.', 'danger')
            return redirect(url_for('movimientos_list'))
        if mes_esta_cerrado(db, movimiento['colegio_ref'], movimiento['curso_ref'], movimiento['fecha']) and not current_user.is_admin_global():
            flash('El mes está cerrado para este movimiento. Solo admin puede eliminarlo.', 'danger')
            return redirect(url_for('movimientos_list'))
        db.execute('DELETE FROM pagos_alumnos WHERE movimiento_id = ?', (movimiento_id,))
        db.execute('DELETE FROM movimientos WHERE id = ?', (movimiento_id,))
        db.commit()
        flash(f'Movimiento eliminado: {movimiento["concepto"]}.', 'success')
        next_url = request.form.get('next', '').strip()
        return redirect(next_url or url_for('movimientos_list'))

    @app.route('/actividades')
    @login_required
    def actividades_list():
        db = get_db()
        actividades_sql = """
            SELECT
                a.id,
                a.nombre,
                a.fecha,
                COALESCE(a.curso, '') AS curso,
                COALESCE(a.descripcion, '') AS descripcion,
                COALESCE(SUM(CASE WHEN m.tipo = 'ingreso' THEN m.monto ELSE 0 END), 0) AS ingresos,
                COALESCE(SUM(CASE WHEN m.tipo = 'gasto' THEN m.monto ELSE 0 END), 0) AS egresos,
                COALESCE(SUM(CASE WHEN m.tipo = 'ingreso' THEN 1 ELSE 0 END), 0) AS cantidad_ingresos,
                COALESCE(SUM(CASE WHEN m.tipo = 'gasto' THEN 1 ELSE 0 END), 0) AS cantidad_egresos
            FROM actividades a
            LEFT JOIN movimientos m ON m.actividad_id = a.id
            WHERE 1=1
        """
        actividades_params: list[Any] = []
        actividades_sql, actividades_params = course_filter_sql(actividades_sql, actividades_params, 'a')
        actividades_sql += ' GROUP BY a.id, a.nombre, a.fecha, a.curso, a.descripcion ORDER BY a.fecha DESC, a.nombre'
        actividades = db.fetchall(actividades_sql, actividades_params)
        resumen_general = {
            'ingresos': sum(float(a['ingresos'] or 0) for a in actividades),
            'egresos': sum(float(a['egresos'] or 0) for a in actividades),
            'cantidad_actividades': len(actividades),
        }
        resumen_general['balance'] = resumen_general['ingresos'] - resumen_general['egresos']
        return render_template('actividades_list.html', actividades=actividades, resumen_general=resumen_general)

    @app.route('/reportes/actividades')
    @login_required
    def actividades_report():
        db = get_db()
        mes = request.args.get('mes') or datetime.today().strftime('%Y-%m')
        actividades_sql = """
            SELECT
                a.id, a.nombre, a.fecha, COALESCE(a.curso, '') AS curso, COALESCE(a.descripcion, '') AS descripcion,
                COALESCE(SUM(CASE WHEN m.tipo = 'ingreso' THEN m.monto ELSE 0 END), 0) AS ingresos,
                COALESCE(SUM(CASE WHEN m.tipo = 'gasto' THEN m.monto ELSE 0 END), 0) AS egresos,
                COUNT(m.id) AS movimientos
            FROM actividades a
            LEFT JOIN movimientos m ON m.actividad_id = a.id
            WHERE 1=1
        """
        actividades_params: list[Any] = []
        actividades_sql, actividades_params = course_filter_sql(actividades_sql, actividades_params, 'a')
        actividades_sql += ' GROUP BY a.id, a.nombre, a.fecha, a.curso, a.descripcion ORDER BY a.fecha DESC, a.nombre'
        actividades = db.fetchall(actividades_sql, actividades_params)
        deudas = resumen_cuotas_por_alumno(db, mes, current_course_filter(), selected_colegio_id() if current_user.is_admin_global() else getattr(current_user, 'colegio_id', None))
        total_deuda = sum(max(float(f['cuota_mensual']) - float(f['pagado']), 0) for f in deudas if f['activo'])
        return render_template('actividades_report.html', actividades=actividades, mes=mes, deudas=deudas, total_deuda=total_deuda)

    @app.route('/actividades/<int:actividad_id>')
    @login_required
    def actividad_detail(actividad_id: int):
        db = get_db()
        actividad = fetch_actividad_permitida(db, actividad_id)
        if not actividad:
            flash('Actividad no encontrada.', 'danger')
            return redirect(url_for('actividades_list'))

        resumen = db.fetchone(
            """
            SELECT
                COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) AS ingresos,
                COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) AS gastos,
                COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN 1 ELSE 0 END), 0) AS cantidad_ingresos,
                COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN 1 ELSE 0 END), 0) AS cantidad_gastos
            FROM movimientos
            WHERE actividad_id = ?
            """
            , (actividad_id,)
        )
        ingresos = db.fetchall(
            """
            SELECT m.id, m.fecha, m.concepto, m.monto, COALESCE(m.observacion, '') AS observacion,
                   COALESCE(m.origen, 'general') AS origen, COALESCE(al.nombre, '-') AS alumno
            FROM movimientos m
            LEFT JOIN alumnos al ON al.id = m.alumno_id
            WHERE m.actividad_id = ? AND m.tipo = 'ingreso'
            ORDER BY m.fecha DESC, m.id DESC
            """
            , (actividad_id,)
        )
        gastos = db.fetchall(
            """
            SELECT m.id, m.fecha, m.concepto, m.monto, COALESCE(m.observacion, '') AS observacion,
                   COALESCE(m.origen, 'general') AS origen
            FROM movimientos m
            WHERE m.actividad_id = ? AND m.tipo = 'gasto'
            ORDER BY m.fecha DESC, m.id DESC
            """
            , (actividad_id,)
        )
        return render_template('actividad_detail.html', actividad=actividad, resumen=resumen, ingresos=ingresos, gastos=gastos)

    @app.route('/actividades/nueva', methods=['GET', 'POST'])
    @role_required('admin', 'presidente', 'tesorero', 'secretario')
    def actividades_new():
        db = get_db()
        if request.method == 'POST':
            nombre = request.form.get('nombre', '').strip()
            fecha = request.form.get('fecha', '').strip()
            curso = (request.form.get('curso', '').strip() or user_course_scope() or '').strip()
            descripcion = request.form.get('descripcion', '').strip()
            try:
                validar_fecha(fecha)
            except Exception:
                flash('Fecha inválida.', 'danger')
                return render_template('actividades_form.html', actividad=None)
            if not ensure_course_access(curso, resolve_colegio_for_course(curso)):
                flash('No puedes crear actividades en otro curso.', 'danger')
                return render_template('actividades_form.html', actividad=None)
            db.execute('INSERT INTO actividades (nombre, fecha, curso, colegio_id, descripcion) VALUES (?, ?, ?, ?, ?)', (nombre, fecha, curso or None, resolve_colegio_for_course(curso), descripcion))
            db.commit()
            flash('Actividad creada.', 'success')
            return redirect(url_for('actividades_list'))
        return render_template('actividades_form.html', actividad=None)

    @app.route('/actividades/<int:actividad_id>/editar', methods=['GET', 'POST'])
    @role_required('admin', 'presidente', 'tesorero', 'secretario')
    def actividades_edit(actividad_id: int):
        db = get_db()
        actividad = fetch_actividad_permitida(db, actividad_id)
        if not actividad:
            flash('Actividad no encontrada.', 'danger')
            return redirect(url_for('actividades_list'))
        if request.method == 'POST':
            nombre = request.form.get('nombre', '').strip()
            fecha = request.form.get('fecha', '').strip()
            curso = (request.form.get('curso', '').strip() or user_course_scope() or '').strip()
            descripcion = request.form.get('descripcion', '').strip()
            try:
                validar_fecha(fecha)
            except Exception:
                flash('Fecha inválida.', 'danger')
                return render_template('actividades_form.html', actividad=actividad)
            if not ensure_course_access(curso, resolve_colegio_for_course(curso)):
                flash('No puedes mover la actividad a otro curso.', 'danger')
                return render_template('actividades_form.html', actividad=actividad)
            db.execute('UPDATE actividades SET nombre=?, fecha=?, curso=?, colegio_id=?, descripcion=? WHERE id=?', (nombre, fecha, curso or None, resolve_colegio_for_course(curso), descripcion, actividad_id))
            db.commit()
            flash('Actividad actualizada.', 'success')
            return redirect(url_for('actividades_list'))
        return render_template('actividades_form.html', actividad=actividad)

    @app.post('/actividades/<int:actividad_id>/eliminar')
    @role_required('admin', 'presidente', 'tesorero', 'secretario')
    def actividades_delete(actividad_id: int):
        db = get_db()
        actividad = fetch_actividad_permitida(db, actividad_id)
        if not actividad:
            flash('Actividad no encontrada.', 'danger')
            return redirect(url_for('actividades_list'))
        db.execute('UPDATE movimientos SET actividad_id = NULL WHERE actividad_id = ?', (actividad_id,))
        db.execute('DELETE FROM actividades WHERE id = ?', (actividad_id,))
        db.commit()
        flash(f'Actividad eliminada: {actividad["nombre"]}.', 'success')
        return redirect(url_for('actividades_list'))


    @app.route('/cuotas/configuracion', methods=['GET', 'POST'])
    @role_required('admin', 'presidente', 'tesorero')
    def cuotas_configuracion():
        db = get_db()
        mes = request.values.get('mes', datetime.now().strftime('%Y-%m')).strip()
        colegio_id = selected_colegio_id() or (int(getattr(current_user, 'colegio_id', 0) or 0) if not current_user.is_admin_global() else None)
        curso = request.values.get('curso', '').strip() or current_course_filter()
        if request.method == 'POST':
            raw_colegio = request.form.get('colegio_id', '').strip()
            if raw_colegio.isdigit():
                colegio_id = int(raw_colegio)
            curso = request.form.get('curso', '').strip()
            monto = parse_float(request.form.get('monto', '0'))
            aplicar_alumnos = request.form.get('aplicar_alumnos') == '1'
            try:
                datetime.strptime(mes + '-01', '%Y-%m-%d')
            except Exception:
                flash('Mes inválido.', 'danger')
                return redirect(url_for('cuotas_configuracion'))
            if not colegio_id or not curso:
                flash('Debes seleccionar colegio y curso.', 'danger')
            elif not ensure_course_access(curso, colegio_id):
                flash('No tienes permisos para ese colegio/curso.', 'danger')
            elif monto < 0:
                flash('La cuota no puede ser negativa.', 'danger')
            elif mes_esta_cerrado(db, colegio_id, curso, mes) and not current_user.is_admin_global():
                flash('El mes está cerrado. Solo admin puede modificar cuota.', 'danger')
            else:
                db.execute("""
                    INSERT INTO cuotas_mensuales (colegio_id, curso, mes, monto, creado_por, creado_en)
                    VALUES (?, ?, ?, ?, ?, ?)
                    ON CONFLICT(colegio_id, curso, mes) DO UPDATE SET monto=excluded.monto, creado_por=excluded.creado_por, creado_en=excluded.creado_en
                """, (colegio_id, curso, mes, monto, current_user.username, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                if aplicar_alumnos:
                    db.execute('UPDATE alumnos SET cuota_mensual = ? WHERE colegio_id = ? AND lower(trim(curso)) = lower(trim(?)) AND activo = 1', (monto, colegio_id, curso))
                db.commit()
                log_audit('configurar', 'cuota_mensual', f'{colegio_id}-{curso}-{mes}', f'Monto {monto}; aplicar_alumnos={aplicar_alumnos}')
                flash('Cuota mensual configurada.', 'success')
                return redirect(url_for('cuotas_configuracion', colegio_id=colegio_id, curso=curso, mes=mes))
        sql = """
            SELECT cm.*, c.nombre AS colegio_nombre
            FROM cuotas_mensuales cm
            LEFT JOIN colegios c ON c.id = cm.colegio_id
            WHERE 1=1
        """
        params=[]
        sql, params = append_scope_filter(sql, params, 'cm.curso', 'cm.colegio_id')
        if mes:
            sql += ' AND cm.mes = ?'
            params.append(mes)
        sql += ' ORDER BY c.nombre, cm.curso, cm.mes DESC'
        cuotas_cfg = db.fetchall(sql, params)
        return render_template('cuotas_configuracion.html', cuotas_cfg=cuotas_cfg, mes=mes, curso=curso, colegio_id=colegio_id)

    @app.route('/notificaciones')
    @login_required
    def notificaciones_view():
        db = get_db()
        mes = request.args.get('mes', datetime.now().strftime('%Y-%m')).strip()
        filas = resumen_deuda_acumulada_por_alumno(db, mes, current_course_filter())
        deudores = [f for f in filas if f.get('activo') and float(f.get('deuda_total') or 0) > 0]
        alumnos_ids = [int(f['id']) for f in deudores]
        contacto_map = {}
        if alumnos_ids:
            placeholders = ','.join(['?'] * len(alumnos_ids))
            for row in db.fetchall(f'SELECT id, apoderado, telefono, email FROM alumnos WHERE id IN ({placeholders})', alumnos_ids):
                contacto_map[int(row['id'])] = row
        mensajes=[]
        for fila in deudores:
            contacto = contacto_map.get(int(fila['id']))
            apoderado = (contacto['apoderado'] if contacto else '') or 'apoderado/a'
            telefono = (contacto['telefono'] if contacto else '') or ''
            mensajes.append({
                'alumno': fila['nombre'],
                'curso': fila['curso'],
                'telefono': telefono,
                'email': ((contacto['email'] if contacto and 'email' in contacto.keys() else '') or ''),
                'alumno_id': fila['id'],
                'deuda': fila['deuda_total'],
                'mensaje': f'Estimado/a {apoderado}, el alumno/a {fila["nombre"]} del curso {fila["curso"]} mantiene una deuda de {formato_monto(fila["deuda_total"])} al mes {mes}. Favor regularizar con tesorería. Gracias.'
            })
        return render_template('notificaciones.html', mensajes=mensajes, mes=mes)

    def enviar_correo_smtp(destinatario: str, asunto: str, cuerpo: str, adjuntos: list[tuple[str, bytes, str]] | None = None) -> tuple[bool, str]:
        host = os.getenv('SMTP_HOST', '').strip()
        port = int(os.getenv('SMTP_PORT', '587') or 587)
        user = os.getenv('SMTP_USER', '').strip()
        password = os.getenv('SMTP_PASSWORD', '').strip()
        sender = os.getenv('SMTP_FROM', user or '').strip()
        if not host or not sender:
            return False, 'SMTP no configurado. Define SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD y SMTP_FROM.'
        msg = EmailMessage()
        msg['Subject'] = asunto
        msg['From'] = sender
        msg['To'] = destinatario
        msg.set_content(cuerpo)
        for nombre_archivo, contenido, mime_type in (adjuntos or []):
            maintype, subtype = (mime_type.split('/', 1) + ['octet-stream'])[:2] if '/' in mime_type else ('application', 'octet-stream')
            msg.add_attachment(contenido, maintype=maintype, subtype=subtype, filename=nombre_archivo)
        try:
            with smtplib.SMTP(host, port, timeout=15) as smtp:
                smtp.starttls()
                if user and password:
                    smtp.login(user, password)
                smtp.send_message(msg)
            return True, 'Correo enviado correctamente.'
        except Exception as exc:
            return False, f'No se pudo enviar el correo: {exc}'

    @app.post('/notificaciones/enviar-correo/<int:alumno_id>')
    @login_required
    @role_required('admin', 'presidente', 'tesorero', 'secretario')
    def notificaciones_enviar_correo(alumno_id: int):
        db = get_db()
        alumno = fetch_alumno_permitido(db, alumno_id)
        if not alumno:
            flash('Alumno no encontrado o sin permisos.', 'danger')
            return redirect(url_for('notificaciones_view'))
        email_destino = (alumno['email'] if 'email' in alumno.keys() else '') or ''
        if not email_destino:
            flash('El alumno no tiene correo guardado en su ficha.', 'warning')
            return redirect(request.referrer or url_for('notificaciones_view'))
        asunto = request.form.get('asunto', '').strip() or f'Notificación de deuda - {APP_NAME}'
        cuerpo = request.form.get('mensaje', '').strip()
        if not cuerpo:
            flash('No hay mensaje para enviar.', 'warning')
            return redirect(request.referrer or url_for('notificaciones_view'))
        ok, msg = enviar_correo_smtp(email_destino, asunto, cuerpo)
        flash(msg, 'success' if ok else 'danger')
        if ok:
            log_audit('enviar_correo', 'notificacion', alumno_id, f'Correo enviado a {email_destino}')
        return redirect(request.referrer or url_for('notificaciones_view'))

    @app.route('/cuotas')
    @login_required
    def cuotas_view():
        db = get_db()
        mes = request.args.get('mes') or datetime.today().strftime('%Y-%m')
        exportar = request.args.get('exportar', '').strip().lower()
        filtro_reporte = request.args.get('filtro_reporte', 'deuda').strip().lower()
        if filtro_reporte not in ('deuda', 'todos'):
            filtro_reporte = 'deuda'

        if exportar == 'pdf':
            pdf_buffer = construir_pdf_deudores(db, mes, filtro_reporte, current_course_filter())
            filename = f'reporte_cuotas_{filtro_reporte}_{mes}.pdf'
            return send_file(pdf_buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)

        filas = resumen_cuotas_por_alumno(db, mes, current_course_filter(), selected_colegio_id() if current_user.is_admin_global() else getattr(current_user, 'colegio_id', None))
        if filtro_reporte == 'deuda':
            filas = [
                fila for fila in filas
                if fila['activo'] and max(float(fila['cuota_mensual']) - float(fila['pagado']), 0) > 0
            ]

        total_esperado = sum(float(x['cuota_mensual']) for x in filas if x['activo'])
        total_pagado = sum(float(x['pagado']) for x in filas)
        total_debe = sum(max(float(x['cuota_mensual']) - float(x['pagado']), 0) for x in filas if x['activo'])
        alertas = obtener_alertas_morosidad(db, mes, current_course_filter(), selected_colegio_id() if current_user.is_admin_global() else getattr(current_user, 'colegio_id', None))
        if filtro_reporte == 'deuda':
            alertas = [alerta for alerta in alertas if alerta['debe'] > 0]
        return render_template(
            'cuotas.html',
            filas=filas,
            mes=mes,
            total_esperado=total_esperado,
            total_pagado=total_pagado,
            total_debe=total_debe,
            alertas=alertas,
            filtro_reporte=filtro_reporte,
            page_title='Estado de cuotas',
            is_morosidad=False,
        )

    @app.route('/morosidad')
    @login_required
    def morosidad_view():
        db = get_db()
        mes = request.args.get('mes') or datetime.today().strftime('%Y-%m')
        # Morosidad siempre muestra solo alumnos con deuda.
        filtro_reporte = 'deuda'
        filas = resumen_cuotas_por_alumno(
            db,
            mes,
            current_course_filter(),
            selected_colegio_id() if current_user.is_admin_global() else getattr(current_user, 'colegio_id', None),
        )
        filas = [
            fila for fila in filas
            if fila['activo'] and max(float(fila['cuota_mensual']) - float(fila['pagado']), 0) > 0
        ]
        total_esperado = sum(float(x['cuota_mensual']) for x in filas if x['activo'])
        total_pagado = sum(float(x['pagado']) for x in filas)
        total_debe = sum(max(float(x['cuota_mensual']) - float(x['pagado']), 0) for x in filas if x['activo'])
        alertas = obtener_alertas_morosidad(
            db,
            mes,
            current_course_filter(),
            selected_colegio_id() if current_user.is_admin_global() else getattr(current_user, 'colegio_id', None),
        )
        alertas = [alerta for alerta in alertas if alerta['debe'] > 0]
        return render_template(
            'cuotas.html',
            filas=filas,
            mes=mes,
            total_esperado=total_esperado,
            total_pagado=total_pagado,
            total_debe=total_debe,
            alertas=alertas,
            filtro_reporte=filtro_reporte,
            page_title='Morosidad',
            is_morosidad=True,
        )

    @app.route('/cuotas/exportar/excel')
    @login_required
    def cuotas_export_excel():
        db = get_db()
        mes = request.args.get('mes') or datetime.today().strftime('%Y-%m')
        filtro_reporte = request.args.get('filtro_reporte', 'deuda').strip().lower()
        filas = resumen_cuotas_por_alumno(db, mes, current_course_filter(), selected_colegio_id() if current_user.is_admin_global() else getattr(current_user, 'colegio_id', None))
        if filtro_reporte == 'deuda':
            filas = [f for f in filas if f['activo'] and max(float(f['cuota_mensual']) - float(f['pagado']), 0) > 0]
        wb = Workbook()
        ws = wb.active
        ws.title = 'Estado cuotas'
        ws.append(['Colegio', 'Alumno', 'Curso', 'Mes', 'Cuota', 'Pagado', 'Debe', 'Estado', 'Activo'])
        for f in filas:
            debe = max(float(f['cuota_mensual']) - float(f['pagado']), 0)
            estado, _ = estado_cuota(f['cuota_mensual'], f['pagado'])
            ws.append([f['colegio_nombre'] if 'colegio_nombre' in f.keys() else '', f['nombre'], f['curso'] or '', mes, float(f['cuota_mensual']), float(f['pagado']), debe, estado, 'Sí' if f['activo'] else 'No'])
        for col in ws.columns:
            width = max(len(str(cell.value or '')) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(width, 45)
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(bio, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'estado_cuotas_{mes}.xlsx')

    def construir_comprobante_pago_pdf(db: DBAdapter, pago_id: int) -> tuple[bytes, str, str, str]:
        """Genera el PDF del comprobante y retorna: bytes, folio, email destino, alumno."""
        pago = db.fetchone(
            """
            SELECT p.*, a.nombre AS nombre, a.curso AS curso, COALESCE(a.colegio_id, 1) AS colegio_id,
                   COALESCE(a.email, '') AS email, c.nombre AS colegio_nombre
            FROM pagos_alumnos p
            INNER JOIN alumnos a ON a.id = p.alumno_id
            LEFT JOIN colegios c ON c.id = COALESCE(a.colegio_id, 1)
            WHERE p.id = ?
            """,
            (pago_id,),
        )
        if not pago:
            raise ValueError('Pago no encontrado.')
        colegio_id = pago['colegio_id'] if 'colegio_id' in pago.keys() else 1
        colegio = db.fetchone('SELECT nombre, ubicacion FROM colegios WHERE id = ?', (colegio_id,))
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=18*mm, leftMargin=18*mm, topMargin=18*mm, bottomMargin=18*mm)
        styles = getSampleStyleSheet()
        elements = []
        folio_num = pago['folio'] if 'folio' in pago.keys() and pago['folio'] else pago_id
        folio = f'CC-{int(folio_num):06d}'
        colegio_nombre = colegio['nombre'] if colegio else (pago['colegio_nombre'] if 'colegio_nombre' in pago.keys() else APP_NAME)
        elements.append(Paragraph(APP_NAME, styles['Title']))
        elements.append(Paragraph('Comprobante de pago', styles['Heading2']))
        elements.append(Spacer(1, 8))
        data = [
            ['Folio', folio],
            ['Colegio', colegio_nombre],
            ['Alumno', pago['nombre']],
            ['Curso', pago['curso'] or '-'],
            ['Fecha de pago', pago['fecha']],
            ['Mes pagado', pago['mes']],
            ['Monto', formato_monto(pago['monto'])],
            ['Observación', pago['observacion'] or '-'],
            ['Emitido por', current_user.nombre],
            ['Fecha emisión', datetime.now().strftime('%Y-%m-%d %H:%M')],
        ]
        table = Table(data, colWidths=[42*mm, 120*mm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e0f2fe')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#0f172a')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 18))
        elements.append(Paragraph('Documento generado automáticamente por ContaCurso.', styles['Normal']))
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue(), folio, (pago['email'] if 'email' in pago.keys() else '') or '', pago['nombre']

    @app.route('/pagos/<int:pago_id>/comprobante.pdf')
    @login_required
    def pago_comprobante(pago_id: int):
        db = get_db()
        pago = fetch_pago_permitido(db, pago_id)
        if not pago:
            flash('Pago no encontrado o sin permiso.', 'danger')
            return redirect(url_for('pagos_list'))
        pdf_bytes, folio, _, _ = construir_comprobante_pago_pdf(db, pago_id)
        buffer = BytesIO(pdf_bytes)
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f'comprobante_{folio}.pdf')

    @app.post('/pagos/<int:pago_id>/enviar-comprobante')
    @login_required
    @role_required('admin', 'presidente', 'tesorero', 'secretario')
    def pago_enviar_comprobante(pago_id: int):
        db = get_db()
        pago = fetch_pago_permitido(db, pago_id)
        if not pago:
            flash('Pago no encontrado o sin permiso.', 'danger')
            return redirect(url_for('pagos_list'))
        pdf_bytes, folio, email_destino, alumno_nombre = construir_comprobante_pago_pdf(db, pago_id)
        if not email_destino:
            flash('El alumno no tiene correo guardado en su ficha.', 'warning')
            return redirect(request.referrer or url_for('pagos_list'))
        asunto = f'Comprobante de pago {folio} - {APP_NAME}'
        cuerpo = (
            f'Estimado/a,\n\n'
            f'Adjuntamos el comprobante de pago {folio} correspondiente al alumno/a {alumno_nombre}.\n\n'
            f'Este correo fue generado automáticamente por {APP_NAME}.\n'
        )
        ok, msg = enviar_correo_smtp(
            email_destino,
            asunto,
            cuerpo,
            adjuntos=[(f'comprobante_{folio}.pdf', pdf_bytes, 'application/pdf')],
        )
        flash(msg, 'success' if ok else 'danger')
        if ok:
            log_audit('enviar_correo', 'comprobante_pago', pago_id, f'Comprobante {folio} enviado a {email_destino}')
        return redirect(request.referrer or url_for('pagos_list'))

    @app.errorhandler(500)
    def internal_server_error(exc):
        app.logger.exception('Internal server error: %s', exc)
        flash('Ocurrió un error interno. Revisa los filtros de búsqueda e inténtalo nuevamente.', 'danger')
        destino = request.referrer or (url_for('dashboard') if current_user.is_authenticated else url_for('login'))
        return redirect(destino)

    return app


def is_postgres_url(url: str) -> bool:
    return url.startswith('postgresql://') or url.startswith('postgres://')


def sql_like_ci(value: str) -> str:
    return f"%{(value or '').strip().lower()}%"


def obtener_movimientos_filtrados(db: DBAdapter, tipo: str = 'Todos', mes: str = '', q: str = '', fecha_desde: str = '', fecha_hasta: str = '', actividad_id: str | int = '', alumno_id: str = '', curso_scope: str | None = None, colegio_scope: int | None = None):
    sql = """
        SELECT m.id, m.fecha, m.tipo, m.concepto, m.monto, COALESCE(a.nombre, '-') AS actividad,
               COALESCE(al.nombre, '-') AS alumno,
               COALESCE(m.origen, 'general') AS origen, COALESCE(m.observacion, '') AS observacion
        FROM movimientos m
        LEFT JOIN actividades a ON a.id = m.actividad_id
        LEFT JOIN alumnos al ON al.id = m.alumno_id
        WHERE 1=1
    """
    params: list[Any] = []
    if tipo in ('ingreso', 'gasto'):
        sql += ' AND m.tipo = ?'
        params.append(tipo)
    if mes:
        sql += ' AND substr(m.fecha, 1, 7) = ?'
        params.append(mes)
    if fecha_desde:
        sql += ' AND m.fecha >= ?'
        params.append(fecha_desde)
    if fecha_hasta:
        sql += ' AND m.fecha <= ?'
        params.append(fecha_hasta)
    if actividad_id:
        sql += ' AND m.actividad_id = ?'
        params.append(int(actividad_id))
    if alumno_id:
        sql += ' AND m.alumno_id = ?'
        params.append(int(alumno_id))
    if colegio_scope:
        sql += " AND COALESCE(m.colegio_id, a.colegio_id, al.colegio_id, 1) = ?"
        params.append(int(colegio_scope))
    if curso_scope:
        sql += " AND lower(trim(COALESCE(m.curso, a.curso, al.curso, ''))) = lower(trim(?))"
        params.append(curso_scope)
    if q:
        like = sql_like_ci(q)
        sql += " AND (LOWER(COALESCE(m.concepto, '')) LIKE ? OR LOWER(COALESCE(m.observacion, '')) LIKE ? OR LOWER(COALESCE(m.fecha, '')) LIKE ? OR LOWER(COALESCE(m.origen, '')) LIKE ? OR LOWER(COALESCE(a.nombre, '')) LIKE ? OR LOWER(COALESCE(al.nombre, '')) LIKE ? OR LOWER(COALESCE(al.curso, '')) LIKE ?)"
        params.extend([like, like, like, like, like, like, like])
    sql += ' ORDER BY m.fecha DESC, m.id DESC'
    return db.fetchall(sql, params)


def exportar_movimientos_pdf(movimientos, school_name: str, school_location: str, filtros: dict[str, str]) -> BytesIO:
    data = BytesIO()
    doc = SimpleDocTemplate(data, pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    styles = getSampleStyleSheet()
    elems = []
    elems.append(Paragraph(f'{school_name} · {school_location}', styles['Title']))
    elems.append(Paragraph(f'Reporte de movimientos · generado {datetime.now().strftime("%Y-%m-%d %H:%M")}', styles['Normal']))
    filtros_txt = ' · '.join([f'{k}: {v}' for k, v in filtros.items()])
    elems.append(Paragraph(filtros_txt, styles['Normal']))
    elems.append(Spacer(1, 6))
    total_ing = sum(float(r['monto']) for r in movimientos if r['tipo'] == 'ingreso')
    total_gas = sum(float(r['monto']) for r in movimientos if r['tipo'] == 'gasto')
    elems.append(Paragraph(f'Registros: {len(movimientos)} · Ingresos: {formato_monto(total_ing)} · Gastos: {formato_monto(total_gas)} · Balance: {formato_monto(total_ing-total_gas)}', styles['Heading3']))
    table_data = [['Fecha', 'Tipo', 'Concepto', 'Actividad', 'Alumno', 'Origen', 'Monto']]
    for row in movimientos:
        concepto = str(row['concepto'])
        if len(concepto) > 38:
            concepto = concepto[:35] + '...'
        table_data.append([row['fecha'], row['tipo'], concepto, row['actividad'], row.get('alumno', '-'), row['origen'], formato_monto(row['monto'])])
    table = Table(table_data, repeatRows=1, colWidths=[24*mm, 20*mm, 78*mm, 42*mm, 42*mm, 30*mm, 24*mm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#dfe7ff')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.black),
        ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#cbd5e1')),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (-1,1), (-1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
    ]))
    elems.append(table)
    doc.build(elems)
    data.seek(0)
    return data


def ruta_fuera_de_backups(ruta: Path) -> bool:
    try:
        return ruta.resolve().parent != BACKUP_DIR.resolve()
    except Exception:
        return True

def crear_backup_db(database_path: str | Path) -> Path:
    database_path = str(database_path)
    BACKUP_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if is_postgres_url(database_path):
        destino = BACKUP_DIR / f'backup_contabilidad_{timestamp}.sql'
        parsed = urlparse(database_path)
        env = os.environ.copy()
        if parsed.password:
            env['PGPASSWORD'] = parsed.password
        cmd = [
            'pg_dump',
            '-h', parsed.hostname or 'localhost',
            '-p', str(parsed.port or 5432),
            '-U', parsed.username or 'postgres',
            '-d', (parsed.path or '/').lstrip('/'),
            '-f', str(destino),
        ]
        try:
            subprocess.run(cmd, check=True, env=env, capture_output=True)
        except FileNotFoundError as exc:
            raise RuntimeError('pg_dump no está instalado o no está en el PATH.') from exc
        except subprocess.CalledProcessError as exc:
            raise RuntimeError(exc.stderr.decode('utf-8', errors='ignore') or 'pg_dump falló.') from exc
        return destino
    origen = Path(database_path)
    destino = BACKUP_DIR / f'backup_contabilidad_{timestamp}.db'
    shutil.copy2(origen, destino)
    return destino


def listar_backups():
    if not BACKUP_DIR.exists():
        return []
    archivos = sorted([*BACKUP_DIR.glob('*.db'), *BACKUP_DIR.glob('*.sql')], key=lambda p: p.stat().st_mtime, reverse=True)
    return [
        {
            'nombre': p.name,
            'tamano': p.stat().st_size,
            'modificado': datetime.fromtimestamp(p.stat().st_mtime),
        }
        for p in archivos
    ]


def formato_monto(valor: Any) -> str:
    try:
        valor = float(valor or 0)
    except Exception:
        valor = 0
    return f"${valor:,.0f}".replace(',', '.')


def parse_float(raw: str) -> float:
    txt = (raw or '').strip()
    if not txt:
        return 0.0
    txt = txt.replace('.', '').replace(',', '.') if ',' in txt else txt.replace(',', '.')
    return float(txt)


def validar_fecha(fecha: str) -> bool:
    datetime.strptime(fecha, '%Y-%m-%d')
    return True


def estado_cuota(cuota: Any, pagado: Any) -> tuple[str, str]:
    cuota_f = float(cuota or 0)
    pagado_f = float(pagado or 0)
    if cuota_f <= 0:
        return ('Sin cuota', '⚪')
    if pagado_f >= cuota_f:
        return ('Pagado', '🟢')
    if pagado_f > 0:
        return ('Parcial', '🟡')
    return ('Deuda', '🔴')


def obtener_nombre_alumno(db: DBAdapter, alumno_id: int) -> str:
    row = db.fetchone('SELECT nombre FROM alumnos WHERE id=?', (alumno_id,))
    return row['nombre'] if row else 'Alumno'


def alumno_duplicado(db: DBAdapter, nombre: str, curso: str, exclude_id: int | None = None, colegio_id: int | None = None) -> bool:
    colegio_id = colegio_id or 1
    sql = "SELECT id FROM alumnos WHERE COALESCE(colegio_id, 1) = ? AND lower(trim(nombre)) = lower(trim(?)) AND lower(trim(COALESCE(curso, ''))) = lower(trim(?))"
    params: list[Any] = [colegio_id, nombre, curso or '']
    if exclude_id:
        sql += ' AND id <> ?'
        params.append(exclude_id)
    return db.fetchone(sql, params) is not None


def pago_duplicado(db: DBAdapter, alumno_id: int, mes: str) -> bool:
    row = db.fetchone('SELECT id FROM pagos_alumnos WHERE alumno_id = ? AND mes = ?', (alumno_id, mes))
    return row is not None



def siguiente_folio_pago(db: DBAdapter, colegio_id: int) -> int:
    try:
        row = db.fetchone("SELECT COALESCE(MAX(p.folio), 0) AS ultimo FROM pagos_alumnos p INNER JOIN alumnos a ON a.id = p.alumno_id WHERE COALESCE(a.colegio_id, 1) = ?", (colegio_id,))
        return int(row['ultimo'] or 0) + 1
    except Exception:
        db.rollback()
        return 1

def registrar_pago_alumno(db: DBAdapter, alumno_id: int, fecha: str, mes: str, monto: float, observacion: str, actividad_id: int | None = None, tipo_pago: str = 'cuota_mensual') -> None:
    alumno = db.fetchone('SELECT * FROM alumnos WHERE id = ?', (alumno_id,))
    if not alumno:
        raise ValueError('Alumno no encontrado')
    if tipo_pago == 'cuota_mensual':
        concepto = f'Cuota mensual alumno: {alumno["nombre"]} ({mes})'
        curso = alumno['curso']
        if db.kind == 'postgres':
            cur = db.execute(
                'INSERT INTO movimientos (fecha, tipo, concepto, monto, actividad_id, alumno_id, observacion, origen, curso, colegio_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?) RETURNING id',
                (fecha, 'ingreso', concepto, monto, None, alumno_id, observacion, 'cuota_mensual', curso, alumno['colegio_id'] if 'colegio_id' in alumno.keys() else 1),
            )
            movimiento_id = cur.fetchone()['id']
        else:
            cur = db.execute(
                'INSERT INTO movimientos (fecha, tipo, concepto, monto, actividad_id, alumno_id, observacion, origen, curso, colegio_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                (fecha, 'ingreso', concepto, monto, None, alumno_id, observacion, 'cuota_mensual', curso, alumno['colegio_id'] if 'colegio_id' in alumno.keys() else 1),
            )
            movimiento_id = cur.lastrowid
        colegio_id = alumno['colegio_id'] if 'colegio_id' in alumno.keys() else 1
        folio = siguiente_folio_pago(db, int(colegio_id or 1))
        db.execute(
            'INSERT INTO pagos_alumnos (alumno_id, fecha, mes, monto, observacion, movimiento_id, folio) VALUES (?, ?, ?, ?, ?, ?, ?)',
            (alumno_id, fecha, mes, monto, observacion, movimiento_id, folio),
        )
    else:
        concepto = f'Aporte actividad alumno: {alumno["nombre"]}'
        detalle = observacion if observacion else f'Aporte para actividad registrado en {mes}'
        db.execute(
            'INSERT INTO movimientos (fecha, tipo, concepto, monto, actividad_id, alumno_id, observacion, origen, curso, colegio_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (fecha, 'ingreso', concepto, monto, actividad_id, alumno_id, detalle, 'actividad_alumno', alumno['curso'], alumno['colegio_id'] if 'colegio_id' in alumno.keys() else 1),
        )


def resumen_cuotas_por_alumno(db: DBAdapter, mes: str, curso_scope: str | None = None, colegio_scope: int | None = None):
    sql = """
        SELECT a.id, a.nombre, a.curso, COALESCE(a.colegio_id, 1) AS colegio_id,
               COALESCE(c.nombre, 'Colegio inicial') AS colegio_nombre,
               a.cuota_mensual, a.activo,
               COALESCE(SUM(CASE WHEN p.mes = ? THEN p.monto ELSE 0 END), 0) AS pagado
        FROM alumnos a
        LEFT JOIN colegios c ON c.id = COALESCE(a.colegio_id, 1)
        LEFT JOIN pagos_alumnos p ON a.id = p.alumno_id
        WHERE 1=1
    """
    params: list[Any] = [mes]
    if colegio_scope:
        sql += " AND COALESCE(a.colegio_id, 1) = ?"
        params.append(colegio_scope)
    if curso_scope:
        sql += " AND lower(trim(COALESCE(a.curso, ''))) = lower(trim(?))"
        params.append(curso_scope)
    sql += ' GROUP BY a.id, a.nombre, a.curso, a.colegio_id, c.nombre, a.cuota_mensual, a.activo ORDER BY colegio_nombre, a.curso, a.nombre'
    return db.fetchall(sql, params)


def obtener_alertas_morosidad(db: DBAdapter, mes: str, curso_scope: str | None = None, colegio_scope: int | None = None):
    alertas = []
    for fila in resumen_cuotas_por_alumno(db, mes, curso_scope, colegio_scope):
        if not fila['activo']:
            continue
        debe = max(float(fila['cuota_mensual']) - float(fila['pagado']), 0)
        if debe > 0:
            estado, icono = estado_cuota(fila['cuota_mensual'], fila['pagado'])
            alertas.append({
                'alumno_id': fila['id'],
                'nombre': fila['nombre'],
                'curso': fila['curso'],
                'colegio': fila['colegio_nombre'] if 'colegio_nombre' in fila.keys() else '',
                'debe': debe,
                'estado': estado,
                'icono': icono,
            })
    return alertas


def meses_hasta_corte(mes_corte: str) -> list[str]:
    corte = datetime.strptime(mes_corte + '-01', '%Y-%m-%d')
    mes_inicio = 3
    if corte.month < mes_inicio:
        return []
    return [f"{corte.year}-{mes:02d}" for mes in range(mes_inicio, corte.month + 1)]


def nombre_mes_es(numero_mes: int) -> str:
    nombres = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril', 5: 'mayo', 6: 'junio',
        7: 'julio', 8: 'agosto', 9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre',
    }
    return nombres.get(numero_mes, str(numero_mes))


def resumen_deuda_acumulada_por_alumno(db: DBAdapter, mes_corte: str, curso_scope: str | None = None):
    meses = meses_hasta_corte(mes_corte)
    filas_sql = """
        SELECT a.id, a.nombre, a.curso, a.cuota_mensual, a.activo
        FROM alumnos a
        WHERE 1=1
    """
    filas_params: list[Any] = []
    if curso_scope:
        filas_sql += " AND lower(trim(COALESCE(a.curso, ''))) = lower(trim(?))"
        filas_params.append(curso_scope)
    filas_sql += ' ORDER BY a.curso, a.nombre'
    filas = db.fetchall(filas_sql, filas_params)
    pagos_rows = db.fetchall(
        """
        SELECT p.alumno_id, p.mes, COALESCE(SUM(p.monto), 0) AS monto
        FROM pagos_alumnos p
        WHERE substr(p.mes, 1, 4) = ? AND p.mes <= ?
        GROUP BY p.alumno_id, p.mes
        """,
        (mes_corte[:4], mes_corte),
    )
    pagos_map: dict[tuple[int, str], float] = {}
    for row in pagos_rows:
        pagos_map[(int(row['alumno_id']), row['mes'])] = float(row['monto'] or 0)

    resumen = []
    for fila in filas:
        cuota = float(fila['cuota_mensual'] or 0)
        detalle_deuda = []
        pagado_acumulado = 0.0
        deuda_total = 0.0
        for mes in meses:
            pagado_mes = float(pagos_map.get((int(fila['id']), mes), 0) or 0)
            pagado_acumulado += pagado_mes
            deuda_mes = max(cuota - pagado_mes, 0) if fila['activo'] else 0.0
            deuda_total += deuda_mes
            if fila['activo'] and deuda_mes > 0:
                detalle_deuda.append({
                    'mes': mes,
                    'deuda': deuda_mes,
                })

        esperado_acumulado = cuota * len(meses) if fila['activo'] else 0.0
        resumen.append({
            'id': fila['id'],
            'nombre': fila['nombre'],
            'curso': fila['curso'],
            'cuota_mensual': cuota,
            'activo': fila['activo'],
            'meses_considerados': len(meses),
            'esperado_acumulado': esperado_acumulado,
            'pagado_acumulado': pagado_acumulado,
            'deuda_total': deuda_total,
            'detalle_deuda': detalle_deuda,
        })
    return resumen


def construir_pdf_deudores(db: DBAdapter, mes_corte: str, modo: str = 'deuda', curso_scope: str | None = None) -> BytesIO:
    filas = resumen_deuda_acumulada_por_alumno(db, mes_corte, curso_scope)
    if modo == 'deuda':
        filas = [fila for fila in filas if fila['activo'] and fila['deuda_total'] > 0]
    else:
        filas = [fila for fila in filas if fila['activo']]

    total_esperado = sum(float(f['esperado_acumulado']) for f in filas)
    total_pagado = sum(float(f['pagado_acumulado']) for f in filas)
    total_deuda = sum(float(f['deuda_total']) for f in filas)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    elements = []
    corte_dt = datetime.strptime(mes_corte + '-01', '%Y-%m-%d')
    titulo = 'Reporte de alumnos con deuda' if modo == 'deuda' else 'Reporte general de alumnos'
    filtro_txt = 'solo con deuda' if modo == 'deuda' else 'todos'
    subtitulo = f'Deuda acumulada desde marzo hasta {nombre_mes_es(corte_dt.month)} de {corte_dt.year} · filtro: {filtro_txt}'

    elements.append(Paragraph(f'<b>{SCHOOL_NAME}</b>', styles['Title']))
    elements.append(Paragraph(titulo, styles['Heading2']))
    elements.append(Paragraph(subtitulo, styles['Normal']))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f'Alumnos incluidos: {len(filas)} &nbsp;&nbsp;&nbsp; Total esperado: {formato_monto(total_esperado)} &nbsp;&nbsp;&nbsp; Total pagado: {formato_monto(total_pagado)} &nbsp;&nbsp;&nbsp; Deuda total: {formato_monto(total_deuda)}',
        styles['Normal']
    ))
    elements.append(Spacer(1, 8))

    data = [['Alumno', 'Curso', 'Cuota', 'Meses', 'Esperado', 'Pagado', 'Debe', 'Meses adeudados']]
    for fila in filas:
        meses_adeudados = ', '.join(nombre_mes_es(int(item['mes'][5:7])) for item in fila['detalle_deuda']) or 'Sin deuda'
        data.append([
            fila['nombre'],
            fila['curso'] or '-',
            formato_monto(fila['cuota_mensual']),
            str(fila['meses_considerados']),
            formato_monto(fila['esperado_acumulado']),
            formato_monto(fila['pagado_acumulado']),
            formato_monto(fila['deuda_total']),
            meses_adeudados,
        ])

    table = Table(data, repeatRows=1, colWidths=[60 * mm, 28 * mm, 23 * mm, 15 * mm, 26 * mm, 26 * mm, 24 * mm, 68 * mm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor('#f8fafc')]),
        ('ALIGN', (2, 1), (6, -1), 'RIGHT'),
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def seed_default_admin(db: DBAdapter) -> None:
    try:
        db.execute('INSERT INTO colegios (id, nombre, ubicacion, activo) VALUES (?, ?, ?, 1)', (1, SCHOOL_NAME, SCHOOL_LOCATION))
        db.commit()
    except Exception:
        db.rollback()
    has_user = db.fetchone('SELECT 1 FROM usuarios LIMIT 1')
    if has_user:
        try:
            db.execute("UPDATE usuarios SET role = 'admin' WHERE role = 'admin_global'")
            db.commit()
        except Exception:
            db.rollback()
        return
    username = os.environ.get('ADMIN_USER', 'admin')
    password = os.environ.get('ADMIN_PASSWORD', 'admin123')
    nombre = os.environ.get('ADMIN_NAME', 'Administrador')
    db.execute(
        'INSERT INTO usuarios (username, email, password_hash, role, nombre, curso, colegio_id, activo) VALUES (?, ?, ?, ?, ?, ?, ?, 1)',
        (username, os.environ.get('ADMIN_EMAIL') or None, generate_password_hash(password), 'admin', nombre, None, None),
    )
    db.commit()


def init_db(db: DBAdapter) -> None:
    if db.kind == 'postgres':
        script = """
        CREATE TABLE IF NOT EXISTS colegios (
            id BIGSERIAL PRIMARY KEY,
            nombre TEXT NOT NULL,
            ubicacion TEXT,
            logo_url TEXT,
            color_marca TEXT,
            activo INTEGER NOT NULL DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS actividades (
            id BIGSERIAL PRIMARY KEY,
            nombre TEXT NOT NULL,
            fecha TEXT,
            curso TEXT,
            colegio_id BIGINT NOT NULL DEFAULT 1,
            descripcion TEXT
        );

        CREATE TABLE IF NOT EXISTS movimientos (
            id BIGSERIAL PRIMARY KEY,
            fecha TEXT NOT NULL,
            tipo TEXT NOT NULL CHECK(tipo IN ('ingreso', 'gasto')),
            concepto TEXT NOT NULL,
            monto DOUBLE PRECISION NOT NULL CHECK(monto >= 0),
            actividad_id BIGINT,
            alumno_id BIGINT,
            observacion TEXT,
            origen TEXT NOT NULL DEFAULT 'general',
            curso TEXT,
            colegio_id BIGINT NOT NULL DEFAULT 1,
            CONSTRAINT fk_mov_actividad FOREIGN KEY (actividad_id) REFERENCES actividades(id)
        );

        CREATE TABLE IF NOT EXISTS alumnos (
            id BIGSERIAL PRIMARY KEY,
            nombre TEXT NOT NULL,
            curso TEXT,
            colegio_id BIGINT NOT NULL DEFAULT 1,
            cuota_mensual DOUBLE PRECISION NOT NULL DEFAULT 0,
            apoderado TEXT,
            telefono TEXT,
            email TEXT,
            direccion TEXT,
            observacion_ficha TEXT,
            activo INTEGER NOT NULL DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS pagos_alumnos (
            id BIGSERIAL PRIMARY KEY,
            alumno_id BIGINT NOT NULL,
            fecha TEXT NOT NULL,
            mes TEXT NOT NULL,
            monto DOUBLE PRECISION NOT NULL CHECK(monto >= 0),
            folio BIGINT,
            observacion TEXT,
            movimiento_id BIGINT,
            CONSTRAINT fk_pagos_alumno FOREIGN KEY (alumno_id) REFERENCES alumnos(id),
            CONSTRAINT fk_pagos_mov FOREIGN KEY (movimiento_id) REFERENCES movimientos(id)
        );

        CREATE TABLE IF NOT EXISTS usuarios (
            id BIGSERIAL PRIMARY KEY,
            username TEXT NOT NULL UNIQUE,
            email TEXT,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            nombre TEXT NOT NULL,
            curso TEXT,
            colegio_id BIGINT,
            activo INTEGER NOT NULL DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS usuario_roles_curso (
            id BIGSERIAL PRIMARY KEY,
            usuario_id BIGINT NOT NULL,
            colegio_id BIGINT NOT NULL,
            curso TEXT NOT NULL,
            rol_curso TEXT NOT NULL CHECK(rol_curso IN ('presidente', 'tesorero', 'secretario', 'apoderado')),
            UNIQUE(usuario_id, colegio_id, curso, rol_curso)
        );

        CREATE TABLE IF NOT EXISTS auditoria_acciones (
            id BIGSERIAL PRIMARY KEY,
            fecha TEXT NOT NULL,
            usuario_id BIGINT,
            username TEXT,
            accion TEXT NOT NULL,
            entidad TEXT NOT NULL,
            entidad_id TEXT,
            colegio_id BIGINT,
            curso TEXT,
            detalle TEXT,
            ip TEXT
        );

        CREATE TABLE IF NOT EXISTS cierres_mensuales (
            id BIGSERIAL PRIMARY KEY,
            colegio_id BIGINT NOT NULL,
            curso TEXT NOT NULL,
            mes TEXT NOT NULL,
            ingresos DOUBLE PRECISION NOT NULL DEFAULT 0,
            gastos DOUBLE PRECISION NOT NULL DEFAULT 0,
            saldo DOUBLE PRECISION NOT NULL DEFAULT 0,
            alumnos_activos INTEGER NOT NULL DEFAULT 0,
            total_pagado DOUBLE PRECISION NOT NULL DEFAULT 0,
            deuda_total DOUBLE PRECISION NOT NULL DEFAULT 0,
            creado_por TEXT,
            creado_en TEXT NOT NULL,
            observacion TEXT,
            UNIQUE(colegio_id, curso, mes)
        );

        CREATE TABLE IF NOT EXISTS cuotas_mensuales (
            id BIGSERIAL PRIMARY KEY,
            colegio_id BIGINT NOT NULL,
            curso TEXT NOT NULL,
            mes TEXT NOT NULL,
            monto DOUBLE PRECISION NOT NULL DEFAULT 0,
            creado_por TEXT,
            creado_en TEXT NOT NULL,
            UNIQUE(colegio_id, curso, mes)
        );


        CREATE TABLE IF NOT EXISTS password_reset_tokens (
            id BIGSERIAL PRIMARY KEY,
            usuario_id BIGINT NOT NULL,
            token TEXT NOT NULL UNIQUE,
            expira_en TEXT NOT NULL,
            usado INTEGER NOT NULL DEFAULT 0
        );

        """
    else:
        script = """
        CREATE TABLE IF NOT EXISTS colegios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            ubicacion TEXT,
            logo_url TEXT,
            color_marca TEXT,
            activo INTEGER NOT NULL DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS actividades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            fecha TEXT,
            curso TEXT,
            colegio_id INTEGER NOT NULL DEFAULT 1,
            descripcion TEXT
        );

        CREATE TABLE IF NOT EXISTS movimientos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            tipo TEXT NOT NULL CHECK(tipo IN ('ingreso', 'gasto')),
            concepto TEXT NOT NULL,
            monto REAL NOT NULL CHECK(monto >= 0),
            actividad_id INTEGER,
            alumno_id INTEGER,
            observacion TEXT,
            origen TEXT NOT NULL DEFAULT 'general',
            curso TEXT,
            colegio_id INTEGER NOT NULL DEFAULT 1,
            FOREIGN KEY (actividad_id) REFERENCES actividades(id)
        );

        CREATE TABLE IF NOT EXISTS alumnos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            curso TEXT,
            colegio_id INTEGER NOT NULL DEFAULT 1,
            cuota_mensual REAL NOT NULL DEFAULT 0,
            apoderado TEXT,
            telefono TEXT,
            email TEXT,
            direccion TEXT,
            observacion_ficha TEXT,
            activo INTEGER NOT NULL DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS pagos_alumnos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            alumno_id INTEGER NOT NULL,
            fecha TEXT NOT NULL,
            mes TEXT NOT NULL,
            monto REAL NOT NULL CHECK(monto >= 0),
            folio INTEGER,
            observacion TEXT,
            movimiento_id INTEGER,
            FOREIGN KEY (alumno_id) REFERENCES alumnos(id),
            FOREIGN KEY (movimiento_id) REFERENCES movimientos(id)
        );

        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            email TEXT,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            nombre TEXT NOT NULL,
            curso TEXT,
            colegio_id INTEGER,
            activo INTEGER NOT NULL DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS usuario_roles_curso (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario_id INTEGER NOT NULL,
            colegio_id INTEGER NOT NULL,
            curso TEXT NOT NULL,
            rol_curso TEXT NOT NULL CHECK(rol_curso IN ('presidente', 'tesorero', 'secretario', 'apoderado')),
            UNIQUE(usuario_id, colegio_id, curso, rol_curso)
        );

        CREATE TABLE IF NOT EXISTS auditoria_acciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            usuario_id INTEGER,
            username TEXT,
            accion TEXT NOT NULL,
            entidad TEXT NOT NULL,
            entidad_id TEXT,
            colegio_id INTEGER,
            curso TEXT,
            detalle TEXT,
            ip TEXT
        );

        CREATE TABLE IF NOT EXISTS cierres_mensuales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            colegio_id INTEGER NOT NULL,
            curso TEXT NOT NULL,
            mes TEXT NOT NULL,
            ingresos REAL NOT NULL DEFAULT 0,
            gastos REAL NOT NULL DEFAULT 0,
            saldo REAL NOT NULL DEFAULT 0,
            alumnos_activos INTEGER NOT NULL DEFAULT 0,
            total_pagado REAL NOT NULL DEFAULT 0,
            deuda_total REAL NOT NULL DEFAULT 0,
            creado_por TEXT,
            creado_en TEXT NOT NULL,
            observacion TEXT,
            UNIQUE(colegio_id, curso, mes)
        );

        CREATE TABLE IF NOT EXISTS cuotas_mensuales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            colegio_id INTEGER NOT NULL,
            curso TEXT NOT NULL,
            mes TEXT NOT NULL,
            monto REAL NOT NULL DEFAULT 0,
            creado_por TEXT,
            creado_en TEXT NOT NULL,
            UNIQUE(colegio_id, curso, mes)
        );


        CREATE TABLE IF NOT EXISTS password_reset_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario_id INTEGER NOT NULL,
            token TEXT NOT NULL UNIQUE,
            expira_en TEXT NOT NULL,
            usado INTEGER NOT NULL DEFAULT 0
        );

        """
    db.executescript(script)

    if db.kind == 'postgres':
        migration_statements = [
            'ALTER TABLE colegios ADD COLUMN IF NOT EXISTS logo_url TEXT',
            'ALTER TABLE colegios ADD COLUMN IF NOT EXISTS color_marca TEXT',
            "ALTER TABLE colegios ADD COLUMN IF NOT EXISTS plan TEXT DEFAULT 'basico'",
            "ALTER TABLE colegios ADD COLUMN IF NOT EXISTS estado_suscripcion TEXT DEFAULT 'activo'",
            'ALTER TABLE colegios ADD COLUMN IF NOT EXISTS fecha_vencimiento TEXT',
            'ALTER TABLE actividades ADD COLUMN IF NOT EXISTS curso TEXT',
            'ALTER TABLE actividades ADD COLUMN IF NOT EXISTS colegio_id BIGINT NOT NULL DEFAULT 1',
            'ALTER TABLE actividades ADD COLUMN IF NOT EXISTS descripcion TEXT',
            'ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS actividad_id BIGINT',
            'ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS alumno_id BIGINT',
            "ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS origen TEXT NOT NULL DEFAULT 'general'",
            'ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS curso TEXT',
            'ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS colegio_id BIGINT NOT NULL DEFAULT 1',
            'ALTER TABLE pagos_alumnos ADD COLUMN IF NOT EXISTS observacion TEXT',
            'ALTER TABLE pagos_alumnos ADD COLUMN IF NOT EXISTS folio BIGINT',
            'ALTER TABLE pagos_alumnos ADD COLUMN IF NOT EXISTS movimiento_id BIGINT',
            'ALTER TABLE alumnos ADD COLUMN IF NOT EXISTS colegio_id BIGINT NOT NULL DEFAULT 1',
            'ALTER TABLE alumnos ADD COLUMN IF NOT EXISTS apoderado TEXT',
            'ALTER TABLE alumnos ADD COLUMN IF NOT EXISTS telefono TEXT',
            'ALTER TABLE alumnos ADD COLUMN IF NOT EXISTS email TEXT',
            'ALTER TABLE alumnos ADD COLUMN IF NOT EXISTS direccion TEXT',
            'ALTER TABLE alumnos ADD COLUMN IF NOT EXISTS observacion_ficha TEXT',
            'ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS email TEXT',
            'ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS curso TEXT',
            'ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS colegio_id BIGINT',
        ]
    else:
        migration_statements = [
            'ALTER TABLE colegios ADD COLUMN logo_url TEXT',
            'ALTER TABLE colegios ADD COLUMN color_marca TEXT',
            "ALTER TABLE colegios ADD COLUMN plan TEXT DEFAULT 'basico'",
            "ALTER TABLE colegios ADD COLUMN estado_suscripcion TEXT DEFAULT 'activo'",
            'ALTER TABLE colegios ADD COLUMN fecha_vencimiento TEXT',
            'ALTER TABLE actividades ADD COLUMN curso TEXT',
            'ALTER TABLE actividades ADD COLUMN colegio_id INTEGER NOT NULL DEFAULT 1',
            'ALTER TABLE actividades ADD COLUMN descripcion TEXT',
            'ALTER TABLE movimientos ADD COLUMN actividad_id INTEGER',
            'ALTER TABLE movimientos ADD COLUMN alumno_id INTEGER',
            "ALTER TABLE movimientos ADD COLUMN origen TEXT NOT NULL DEFAULT 'general'",
            'ALTER TABLE movimientos ADD COLUMN curso TEXT',
            'ALTER TABLE movimientos ADD COLUMN colegio_id INTEGER NOT NULL DEFAULT 1',
            'ALTER TABLE pagos_alumnos ADD COLUMN observacion TEXT',
            'ALTER TABLE pagos_alumnos ADD COLUMN folio INTEGER',
            'ALTER TABLE pagos_alumnos ADD COLUMN movimiento_id INTEGER',
            'ALTER TABLE alumnos ADD COLUMN colegio_id INTEGER NOT NULL DEFAULT 1',
            'ALTER TABLE alumnos ADD COLUMN apoderado TEXT',
            'ALTER TABLE alumnos ADD COLUMN telefono TEXT',
            'ALTER TABLE alumnos ADD COLUMN email TEXT',
            'ALTER TABLE alumnos ADD COLUMN direccion TEXT',
            'ALTER TABLE alumnos ADD COLUMN observacion_ficha TEXT',
            'ALTER TABLE usuarios ADD COLUMN email TEXT',
            'ALTER TABLE usuarios ADD COLUMN curso TEXT',
            'ALTER TABLE usuarios ADD COLUMN colegio_id INTEGER',
        ]

    for statement in migration_statements:
        try:
            db.execute(statement)
            db.commit()
        except Exception:
            db.rollback()

    if db.kind == 'sqlite':
        # SQLite no permite modificar un CHECK existente con ALTER TABLE.
        # Se reconstruye usuario_roles_curso para permitir el rol 'apoderado'.
        try:
            db.executescript("""
            DROP TABLE IF EXISTS usuario_roles_curso_legacy;
            ALTER TABLE usuario_roles_curso RENAME TO usuario_roles_curso_legacy;
            CREATE TABLE usuario_roles_curso (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                usuario_id INTEGER NOT NULL,
                colegio_id INTEGER NOT NULL,
                curso TEXT NOT NULL,
                rol_curso TEXT NOT NULL CHECK(rol_curso IN ('presidente', 'tesorero', 'secretario', 'apoderado')),
                UNIQUE(usuario_id, colegio_id, curso, rol_curso)
            );
            INSERT OR IGNORE INTO usuario_roles_curso (id, usuario_id, colegio_id, curso, rol_curso)
            SELECT id, usuario_id, colegio_id, curso, rol_curso FROM usuario_roles_curso_legacy;
            DROP TABLE usuario_roles_curso_legacy;
            """)
            db.commit()
        except Exception:
            db.rollback()

        try:
            db.executescript("""
            DROP TABLE IF EXISTS usuarios_legacy;
            ALTER TABLE usuarios RENAME TO usuarios_legacy;
            CREATE TABLE usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                email TEXT,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL,
                nombre TEXT NOT NULL,
                curso TEXT,
                colegio_id INTEGER,
                activo INTEGER NOT NULL DEFAULT 1
            );
            INSERT INTO usuarios (id, username, email, password_hash, role, nombre, curso, colegio_id, activo)
            SELECT id, username, email, password_hash, role, nombre, curso, colegio_id, activo FROM usuarios_legacy;
            DROP TABLE usuarios_legacy;
            """)
            db.commit()
        except Exception:
            db.rollback()
    else:
        try:
            db.execute('ALTER TABLE usuarios DROP CONSTRAINT IF EXISTS usuarios_role_check')
            db.commit()
        except Exception:
            db.rollback()
        try:
            db.execute('ALTER TABLE usuario_roles_curso DROP CONSTRAINT IF EXISTS usuario_roles_curso_rol_curso_check')
            db.execute("ALTER TABLE usuario_roles_curso ADD CONSTRAINT usuario_roles_curso_rol_curso_check CHECK (rol_curso IN ('presidente', 'tesorero', 'secretario', 'apoderado'))")
            db.commit()
        except Exception:
            db.rollback()


    # Crear índices después de ejecutar las migraciones. En SQLite, si la tabla ya existía
    # sin colegio_id, crear el índice antes del ALTER TABLE provoca: no such column: colegio_id.
    index_statements = [
        'CREATE INDEX IF NOT EXISTS idx_movimientos_fecha ON movimientos(fecha)',
        'CREATE INDEX IF NOT EXISTS idx_movimientos_colegio_curso ON movimientos(colegio_id, curso)',
        'CREATE INDEX IF NOT EXISTS idx_alumnos_colegio_curso ON alumnos(colegio_id, curso)',
        'CREATE INDEX IF NOT EXISTS idx_actividades_colegio_curso ON actividades(colegio_id, curso)',
        'CREATE INDEX IF NOT EXISTS idx_usuario_roles_curso_lookup ON usuario_roles_curso(usuario_id, colegio_id, curso)',
        'CREATE INDEX IF NOT EXISTS idx_pagos_alumnos_mes ON pagos_alumnos(mes)',
        'CREATE INDEX IF NOT EXISTS idx_auditoria_fecha ON auditoria_acciones(fecha)',
        'CREATE INDEX IF NOT EXISTS idx_auditoria_usuario ON auditoria_acciones(username)',
        'CREATE INDEX IF NOT EXISTS idx_auditoria_entidad ON auditoria_acciones(entidad)',
        'CREATE INDEX IF NOT EXISTS idx_cierres_mes ON cierres_mensuales(mes)',
        'CREATE INDEX IF NOT EXISTS idx_cierres_colegio_curso ON cierres_mensuales(colegio_id, curso)',
        'CREATE INDEX IF NOT EXISTS idx_cuotas_mensuales_lookup ON cuotas_mensuales(colegio_id, curso, mes)',
        'CREATE INDEX IF NOT EXISTS idx_password_reset_token ON password_reset_tokens(token)',
        'CREATE INDEX IF NOT EXISTS idx_pagos_folio ON pagos_alumnos(folio)',
        'CREATE UNIQUE INDEX IF NOT EXISTS idx_pagos_alumno_mes_unique ON pagos_alumnos(alumno_id, mes)',
    ]
    for statement in index_statements:
        try:
            db.execute(statement)
            db.commit()
        except Exception:
            db.rollback()

    try:
        db.execute('DROP INDEX IF EXISTS idx_alumnos_nombre_curso_unique')
        db.commit()
    except Exception:
        db.rollback()
    try:
        if db.kind == 'sqlite':
            db.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_alumnos_colegio_nombre_curso_unique ON alumnos(colegio_id, lower(trim(nombre)), lower(trim(COALESCE(curso, ''))))")
        else:
            db.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_alumnos_colegio_nombre_curso_unique ON alumnos (colegio_id, (lower(trim(nombre))), (lower(trim(COALESCE(curso, '')))))")
        db.commit()
    except Exception:
        db.rollback()

    for statement in [
        "UPDATE usuarios SET role = 'admin' WHERE role = 'admin_global'",
        "UPDATE alumnos SET colegio_id = 1 WHERE colegio_id IS NULL OR colegio_id = 0",
        "UPDATE actividades SET colegio_id = 1 WHERE colegio_id IS NULL OR colegio_id = 0",
        "UPDATE movimientos SET colegio_id = 1 WHERE colegio_id IS NULL OR colegio_id = 0",
        "UPDATE usuarios SET colegio_id = 1 WHERE (colegio_id IS NULL OR colegio_id = 0) AND role NOT IN ('admin', 'admin_global')",
        "UPDATE movimientos SET curso = (SELECT curso FROM alumnos WHERE alumnos.id = movimientos.alumno_id) WHERE (curso IS NULL OR trim(curso) = '') AND alumno_id IS NOT NULL",
        "UPDATE movimientos SET curso = (SELECT curso FROM actividades WHERE actividades.id = movimientos.actividad_id) WHERE (curso IS NULL OR trim(curso) = '') AND actividad_id IS NOT NULL",
        "UPDATE actividades SET curso = (SELECT MIN(m.curso) FROM movimientos m WHERE m.actividad_id = actividades.id AND m.curso IS NOT NULL) WHERE curso IS NULL OR trim(curso) = ''",
        "UPDATE movimientos SET colegio_id = (SELECT colegio_id FROM alumnos WHERE alumnos.id = movimientos.alumno_id) WHERE alumno_id IS NOT NULL AND EXISTS (SELECT 1 FROM alumnos WHERE alumnos.id = movimientos.alumno_id AND alumnos.colegio_id IS NOT NULL AND alumnos.colegio_id <> 0)",
        "UPDATE movimientos SET colegio_id = (SELECT colegio_id FROM actividades WHERE actividades.id = movimientos.actividad_id) WHERE actividad_id IS NOT NULL AND EXISTS (SELECT 1 FROM actividades WHERE actividades.id = movimientos.actividad_id AND actividades.colegio_id IS NOT NULL AND actividades.colegio_id <> 0)",
    ]:
        try:
            db.execute(statement)
            db.commit()
        except Exception:
            db.rollback()


    try:
        pagos_sin_folio = db.fetchall("SELECT p.id, COALESCE(a.colegio_id, 1) AS colegio_id FROM pagos_alumnos p INNER JOIN alumnos a ON a.id = p.alumno_id WHERE p.folio IS NULL OR p.folio = 0 ORDER BY COALESCE(a.colegio_id, 1), p.fecha, p.id")
        folios_por_colegio = {}
        for p in pagos_sin_folio:
            cid = int(p['colegio_id'] or 1)
            if cid not in folios_por_colegio:
                row = db.fetchone("SELECT COALESCE(MAX(p2.folio), 0) AS ultimo FROM pagos_alumnos p2 INNER JOIN alumnos a2 ON a2.id=p2.alumno_id WHERE COALESCE(a2.colegio_id,1)=?", (cid,))
                folios_por_colegio[cid] = int(row['ultimo'] or 0)
            folios_por_colegio[cid] += 1
            db.execute('UPDATE pagos_alumnos SET folio=? WHERE id=?', (folios_por_colegio[cid], p['id']))
        db.commit()
    except Exception:
        db.rollback()

    try:
        legacy_users = db.fetchall("SELECT id, role, curso, colegio_id FROM usuarios WHERE role IN ('presidente', 'tesorero', 'secretario') AND curso IS NOT NULL AND trim(curso) <> ''")
        for u in legacy_users:
            db.execute(
                'INSERT INTO usuario_roles_curso (usuario_id, colegio_id, curso, rol_curso) VALUES (?, ?, ?, ?)',
                (u['id'], u['colegio_id'] or 1, u['curso'], u['role'])
            )
        db.commit()
    except Exception:
        db.rollback()

    normalize_courses_in_db(db)
    db.commit()
app = create_app()

if __name__ == '__main__':
    host = os.environ.get('APP_HOST', '0.0.0.0')
    port = int(os.environ.get('PORT', os.environ.get('APP_PORT', '10000')))
    debug = os.environ.get('APP_DEBUG', '0') == '1'
    app.run(host=host, port=port, debug=debug, use_reloader=False)
